#!/usr/bin/env python3
"""
Streamlit UI for Requirements Extractor from Teams Meetings
Simple web interface for extracting requirements from transcripts.
"""

import streamlit as st
import tempfile
import os
from pathlib import Path
from datetime import datetime
import json
import base64
import io
import threading
import time

# Try to import libraries for PDF and Excel export
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import cryptography for API key encryption
try:
    from cryptography.fernet import Fernet
    CRYPTOGRAPHY_AVAILABLE = True
except ImportError:
    CRYPTOGRAPHY_AVAILABLE = False

from requirements_extractor import (
    TranscriptParser,
    RequirementsExtractor,
    RequirementsFormatter
)

# Import subscription and security modules
try:
    from subscription_manager import SubscriptionManager, SUBSCRIPTION_TIERS
    from security import SecurityManager
    from auth import AuthManager
    SUBSCRIPTION_AVAILABLE = True
    AUTH_AVAILABLE = True
except ImportError as e:
    SUBSCRIPTION_AVAILABLE = False
    AUTH_AVAILABLE = False
    print(f"Warning: Subscription/Auth module not available: {e}")

# Configure FFmpeg path before importing moviepy
try:
    import imageio_ffmpeg
    ffmpeg_path = imageio_ffmpeg.get_ffmpeg_exe()
    os.environ["IMAGEIO_FFMPEG_EXE"] = ffmpeg_path
    os.environ["FFMPEG_BINARY"] = ffmpeg_path
    os.environ["FFMPEG_EXE"] = ffmpeg_path
    # Also set for moviepy
    os.environ["MOVIEPY_FFMPEG_BINARY"] = ffmpeg_path
    
    # Create ffprobe wrapper (ffmpeg can act as ffprobe)
    import os.path as osp
    temp_bin_dir = osp.join(osp.expanduser("~"), ".requirements_extractor_bin")
    os.makedirs(temp_bin_dir, exist_ok=True)
    ffprobe_wrapper = osp.join(temp_bin_dir, "ffprobe")
    
    # Create a shell script that calls ffmpeg with probe-like behavior
    if not osp.exists(ffprobe_wrapper):
        with open(ffprobe_wrapper, 'w') as f:
            f.write(f"""#!/bin/bash
# Wrapper to use ffmpeg as ffprobe
# ffmpeg can handle most probe operations when called correctly
exec "{ffmpeg_path}" -hide_banner -loglevel error "$@"
""")
        os.chmod(ffprobe_wrapper, 0o755)
    
    os.environ["FFPROBE_BINARY"] = ffprobe_wrapper
    os.environ["FFPROBE"] = ffprobe_wrapper
    # Add to PATH so subprocess calls can find it
    os.environ["PATH"] = temp_bin_dir + os.pathsep + os.environ.get("PATH", "")
except Exception as e:
    pass  # If imageio_ffmpeg is not available, try system FFmpeg

# Try to import video processing libraries
try:
    # Try new moviepy 2.x import first
    try:
        from moviepy import VideoFileClip
    except ImportError:
        # Fallback to old moviepy 1.x import
        from moviepy.editor import VideoFileClip
    MOVIEPY_AVAILABLE = True
except ImportError:
    MOVIEPY_AVAILABLE = False

try:
    from pydub import AudioSegment
    PYDUB_AVAILABLE = True
except ImportError:
    PYDUB_AVAILABLE = False

# Try to import local Whisper (no API key needed)
try:
    import whisper
    WHISPER_LOCAL_AVAILABLE = True
except ImportError:
    WHISPER_LOCAL_AVAILABLE = False

# Page configuration with better branding
st.set_page_config(
    page_title="ReqIQ | AI Requirements Extraction",
    page_icon="✨",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': "ReqIQ - AI-powered tool to extract structured requirements from meeting transcripts and videos."
    }
)

# Enhanced Custom CSS for modern, professional styling
st.markdown("""
    <style>
    /* Ensure white background and dark text */
    .stApp {
        background-color: #ffffff !important;
    }
    
    .main .block-container {
        background-color: #ffffff !important;
    }
    
    /* Make all text dark and visible - but exclude input fields */
    h1, h2, h3, h4, h5, h6 {
        color: #000000 !important;
    }
    
    p, span, div, label {
        color: #333333 !important;
    }
    
    /* Force black text on ALL input elements - highest priority, most specific */
    input,
    input[type="text"],
    input[type="password"],
    input[type="email"],
    input[type="number"],
    textarea,
    select,
    .stTextInput input,
    .stTextInput input[type="text"],
    .stTextInput input[type="password"],
    .stTextArea textarea,
    .stSelectbox select,
    .stSelectbox input,
    .stSelectbox > div > div,
    .stSelectbox > div > div > div,
    .stSelectbox > div > div > div input,
    [data-baseweb="select"] input,
    [data-baseweb="input"] input,
    [data-baseweb="textarea"] textarea,
    [data-baseweb="input"] > input,
    [data-baseweb="textarea"] > textarea,
    [data-baseweb="select"] > div > input,
    div[data-baseweb="input"] input,
    div[data-baseweb="textarea"] textarea {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
        background-color: #ffffff !important;
        border: 1px solid #cccccc !important;
    }
    
    /* Force black text with even higher specificity */
    .stTextInput > div > div > div > input,
    .stTextInput > div > div > div > div > input,
    .stTextArea > div > div > div > textarea,
    .stTextArea > div > div > div > div > textarea {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
        background-color: #ffffff !important;
    }
    
    /* Override webkit autofill and any other browser styles */
    input:-webkit-autofill,
    input:-webkit-autofill:hover,
    input:-webkit-autofill:focus,
    input:-webkit-autofill:active {
        -webkit-text-fill-color: #000000 !important;
        color: #000000 !important;
        background-color: #ffffff !important;
    }
    
    /* Input field containers - ensure white background */
    .stTextInput > div > div,
    .stTextInput > div > div > div,
    .stTextArea > div > div,
    .stTextArea > div > div > div,
    .stSelectbox > div > div > div,
    .stSelectbox > div > div > div > div {
        background-color: #ffffff !important;
    }
    
    /* Ensure placeholder text is visible but lighter */
    input::placeholder,
    textarea::placeholder,
    input::-webkit-input-placeholder,
    textarea::-webkit-input-placeholder,
    input::-moz-placeholder,
    textarea::-moz-placeholder {
        color: #666666 !important;
        opacity: 1 !important;
    }
    
    /* Focus state - keep white background and black text */
    .stTextInput input:focus,
    .stTextInput input[type="text"]:focus,
    .stTextInput input[type="password"]:focus,
    .stTextArea textarea:focus,
    .stSelectbox select:focus,
    input[type="text"]:focus,
    input[type="password"]:focus,
    textarea:focus,
    input:focus,
    textarea:focus {
        color: #000000 !important;
        background-color: #ffffff !important;
        border-color: #4a90e2 !important;
    }
    
    /* Override any dark mode styles */
    [data-theme="dark"] input,
    [data-theme="dark"] textarea,
    [data-theme="dark"] select,
    .dark input,
    .dark textarea,
    .dark select {
        color: #000000 !important;
        background-color: #ffffff !important;
    }
    
    /* File uploader - LIGHT background with BLACK text for visibility */
    .stFileUploader,
    [data-testid="stFileUploader"],
    .stFileUploader > div,
    [data-testid="stFileUploader"] > div,
    [data-testid="stFileUploader"] > div > div,
    [data-testid="stFileUploader"] > div > div > div {
        color: #000000 !important;
        background-color: #f8f9fa !important; /* Light gray background */
        border: 2px dashed #cccccc !important;
    }
    
    /* File uploader text visibility - BLACK text on light background */
    .stFileUploader label,
    .stFileUploader p,
    .stFileUploader div,
    .stFileUploader span,
    .uploadedFile,
    [data-testid="stFileUploader"] label,
    [data-testid="stFileUploader"] p,
    [data-testid="stFileUploader"] div,
    [data-testid="stFileUploader"] span,
    [data-testid="stFileUploader"] label p,
    [data-testid="stFileUploader"] label div,
    [data-testid="stFileUploader"] > div > div > div,
    [data-testid="stFileUploader"] > div > div > div > div,
    [data-testid="stFileUploader"] > div > div > div > div > div,
    [data-testid="stFileUploader"] > div > div > div > div > div > div {
        color: #000000 !important;
        background-color: transparent !important;
    }
    
    /* Make ALL text in file uploader BLACK */
    .stFileUploader *,
    [data-testid="stFileUploader"] * {
        color: #000000 !important;
        background-color: transparent !important;
    }
    
    /* Specific targeting for the drag and drop text */
    [data-testid="stFileUploader"] > div > div > div > div > div > div,
    [data-testid="stFileUploader"] > div > div > div > div > div > div > div,
    [data-testid="stFileUploader"] > div > div > div > div > div > div > div > div,
    [data-testid="stFileUploader"] > div > div > div > div > div > div > div > div > div {
        color: #000000 !important;
        background-color: transparent !important;
    }
    
    /* Premium Header Styling */
    .app-title {
        font-size: 3.5rem;
        font-weight: 800;
        color: #000000 !important;
        margin: 0;
        letter-spacing: -0.03em;
        line-height: 1.1;
        text-align: center;
    }
    
    .app-tagline {
        font-size: 1.4rem;
        color: #333333 !important;
        margin: 0.75rem 0 0 0;
        font-weight: 500;
        letter-spacing: 0.01em;
        text-align: center;
    }
    
    .app-subtitle {
        font-size: 1rem;
        color: #555555 !important;
        margin: 0.5rem 0 0 0;
        font-style: italic;
        text-align: center;
    }
    
    /* Legacy header classes for compatibility */
    .main-header {
        font-size: 3.5rem;
        font-weight: 800;
        color: #000000 !important;
        margin-bottom: 0.5rem;
        padding: 0.5rem 0;
        letter-spacing: -0.03em;
    }
    
    .sub-header {
        font-size: 1.3rem;
        color: #333333 !important;
        margin-bottom: 2.5rem;
        font-weight: 500;
        line-height: 1.6;
    }
    
    /* Enhanced Cards and Boxes */
    .success-box {
        padding: 1.25rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #d4edda 0%, #c3e6cb 100%);
        border: 2px solid #28a745;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(40, 167, 69, 0.15);
    }
    
    .info-box {
        padding: 1.25rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #d1ecf1 0%, #bee5eb 100%);
        border: 2px solid #17a2b8;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(23, 162, 184, 0.15);
    }
    
    .warning-box {
        padding: 1.25rem;
        border-radius: 12px;
        background: linear-gradient(135deg, #fff3cd 0%, #ffeaa7 100%);
        border: 2px solid #ffc107;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(255, 193, 7, 0.15);
    }
    
    /* Enhanced Metrics Cards */
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 16px rgba(0, 0, 0, 0.12);
    }
    
    /* Better File Upload Area */
    .uploadedFile {
        border: 2px dashed #e0e0e0;
        border-radius: 12px;
        padding: 2rem;
        background: #ffffff;
        transition: all 0.3s ease;
    }
    
    .uploadedFile:hover {
        border-color: #d0d0d0;
        background: #fafafa;
    }
    
    /* Enhanced Buttons - ensure visibility */
    .stButton > button,
    button,
    [data-baseweb="button"] {
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        color: #ffffff !important;
        background-color: #1f77b4 !important;
        border: 1px solid #1f77b4 !important;
    }
    
    .stButton > button:hover,
    button:hover,
    [data-baseweb="button"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
        background-color: #1565a0 !important;
        border-color: #1565a0 !important;
        color: #ffffff !important;
    }
    
    /* Primary buttons */
    .stButton > button[kind="primary"],
    button[kind="primary"],
    [data-baseweb="button"][kind="primary"] {
        background-color: #1f77b4 !important;
        color: #ffffff !important;
        border-color: #1f77b4 !important;
    }
    
    /* Button text visibility */
    .stButton > button *,
    button *,
    [data-baseweb="button"] * {
        color: #ffffff !important;
    }
    
    /* Sidebar Enhancements - LIGHT background */
    .css-1d391kg,
    [data-testid="stSidebar"],
    section[data-testid="stSidebar"] {
        background: #f8f9fa !important;
    }
    
    /* Sidebar text - BLACK for visibility on light background */
    [data-testid="stSidebar"],
    [data-testid="stSidebar"] *,
    .css-1d391kg,
    .css-1d391kg *,
    .stSidebar,
    .stSidebar *,
    section[data-testid="stSidebar"],
    section[data-testid="stSidebar"] * {
        color: #000000 !important;
    }
    
    /* Sidebar headers and labels - BLACK */
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] h4,
    [data-testid="stSidebar"] h5,
    [data-testid="stSidebar"] h6,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] span {
        color: #000000 !important;
    }
    
    /* Sidebar radio buttons - FORCE BLACK TEXT with maximum specificity */
    [data-testid="stSidebar"] .stRadio,
    [data-testid="stSidebar"] .stRadio *,
    [data-testid="stSidebar"] .stRadio > div,
    [data-testid="stSidebar"] .stRadio > div > div,
    [data-testid="stSidebar"] .stRadio > div > div > label,
    [data-testid="stSidebar"] .stRadio > div > div > label > div,
    [data-testid="stSidebar"] .stRadio > div > div > label > div > div,
    [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div,
    [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div > div,
    [data-testid="stSidebar"] .stRadio label,
    [data-testid="stSidebar"] .stRadio label *,
    [data-testid="stSidebar"] .stRadio span,
    [data-testid="stSidebar"] .stRadio p,
    [data-testid="stSidebar"] .stRadio div,
    [data-testid="stSidebar"] [data-baseweb="radio"],
    [data-testid="stSidebar"] [data-baseweb="radio"] *,
    [data-testid="stSidebar"] [data-baseweb="radio"] label,
    [data-testid="stSidebar"] [data-baseweb="radio"] label *,
    [data-testid="stSidebar"] [data-baseweb="radio"] span,
    [data-testid="stSidebar"] [data-baseweb="radio"] div,
    [data-testid="stSidebar"] [data-baseweb="radio"] p {
        color: #000000 !important;
        background-color: transparent !important;
    }
    
    /* Sidebar other inputs and labels - BLACK */
    [data-testid="stSidebar"] .stTextInput label,
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stInfo,
    [data-testid="stSidebar"] .stMarkdown,
    [data-testid="stSidebar"] .stInfo *,
    [data-testid="stSidebar"] .stMarkdown * {
        color: #000000 !important;
    }
    
    /* Better Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 0.75rem 1.5rem;
        font-weight: 500;
    }
    
    /* Enhanced Expanders */
    .streamlit-expanderHeader {
        font-weight: 600;
        border-radius: 8px;
    }
    
    /* Progress Bar Enhancement */
    .stProgress > div > div > div {
        background: #e0e0e0;
        border-radius: 10px;
    }
    
    /* Better Spacing */
    .main .block-container {
        padding-top: 3rem;
        padding-bottom: 3rem;
    }
    
    /* Status Badges */
    .status-badge {
        display: inline-block;
        padding: 0.35rem 0.75rem;
        border-radius: 20px;
        font-size: 0.875rem;
        font-weight: 600;
        margin: 0.25rem;
    }
    
    .status-success {
        background: #d4edda;
        color: #155724;
        border: 1px solid #c3e6cb;
    }
    
    .status-info {
        background: #d1ecf1;
        color: #0c5460;
        border: 1px solid #bee5eb;
    }
    
    .status-warning {
        background: #fff3cd;
        color: #856404;
        border: 1px solid #ffeaa7;
    }
    
    /* Requirement Card */
    .requirement-card {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 1.25rem;
        margin: 0.75rem 0;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease;
    }
    
    .requirement-card:hover {
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
        border-color: #d0d0d0;
    }
    
    /* Divider Enhancement */
    hr {
        border: none;
        height: 2px;
        background: linear-gradient(90deg, transparent 0%, #e2e8f0 50%, transparent 100%);
        margin: 2rem 0;
    }
    
    /* Input Field Enhancement */
    .stTextInput > div > div > input {
        border-radius: 8px;
        border: 2px solid #e2e8f0;
        transition: all 0.2s ease;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: #d0d0d0;
        box-shadow: 0 0 0 3px rgba(224, 224, 224, 0.3);
    }
    
    /* Selectbox Enhancement */
    .stSelectbox > div > div {
        border-radius: 8px;
    }
    
    /* Radio Button Enhancement */
    .stRadio > div {
        gap: 1rem;
    }
    
    .stRadio > div > label {
        padding: 0.75rem;
        border-radius: 8px;
        border: 2px solid #e2e8f0;
        transition: all 0.2s ease;
    }
    
    .stRadio > div > label:hover {
        border-color: #d0d0d0;
        background: #fafafa;
    }
    
    /* Better Code Blocks */
    .stCodeBlock {
        border-radius: 8px;
        border: 1px solid #e2e8f0;
    }
    
    /* Table Enhancements */
    .dataframe {
        border-radius: 8px;
        overflow: hidden;
    }
    
    /* Hide Streamlit Branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    
    /* Custom Scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: #d0d0d0;
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: #b0b0b0;
    }
    </style>
""", unsafe_allow_html=True)


def get_encryption_key():
    """Get or create encryption key for API key storage."""
    if not CRYPTOGRAPHY_AVAILABLE:
        return None
    try:
        key_file = Path.home() / ".reqiq_key"
        if key_file.exists():
            return key_file.read_bytes()
        else:
            key = Fernet.generate_key()
            key_file.write_bytes(key)
            key_file.chmod(0o600)  # Secure permissions
            return key
    except Exception:
        return None

def save_api_key(api_key: str):
    """Save API key to local config file (encrypted if possible, otherwise base64)."""
    try:
        config_file = Path.home() / ".reqiq_config"
        
        if CRYPTOGRAPHY_AVAILABLE:
            # Use encryption
            key = get_encryption_key()
            if key:
                fernet = Fernet(key)
                encrypted_key = fernet.encrypt(api_key.encode())
                config_file.write_bytes(encrypted_key)
            else:
                # Fallback to base64 encoding
                encoded_key = base64.b64encode(api_key.encode()).decode()
                config_file.write_text(encoded_key)
        else:
            # Fallback to base64 encoding (not secure, but better than plain text)
            encoded_key = base64.b64encode(api_key.encode()).decode()
            config_file.write_text(encoded_key)
        
        config_file.chmod(0o600)  # Secure permissions
        return True
    except Exception as e:
        return False

def load_api_key():
    """Load API key from local config file (decrypted)."""
    try:
        config_file = Path.home() / ".reqiq_config"
        if not config_file.exists():
            return None
        
        if CRYPTOGRAPHY_AVAILABLE:
            # Try decryption first
            try:
                key = get_encryption_key()
                if key:
                    fernet = Fernet(key)
                    encrypted_key = config_file.read_bytes()
                    decrypted_key = fernet.decrypt(encrypted_key).decode()
                    return decrypted_key
            except:
                pass
        
        # Fallback to base64 decoding
        try:
            encoded_key = config_file.read_text()
            decrypted_key = base64.b64decode(encoded_key).decode()
            return decrypted_key
        except:
            return None
    except Exception:
        return None

def get_api_key_from_all_sources():
    """Get API key from all possible sources (priority order)."""
    # Priority 1: Streamlit secrets (for cloud deployment)
    try:
        if hasattr(st, 'secrets') and 'openai_api_key' in st.secrets:
            return st.secrets['openai_api_key']
    except:
        pass
    
    # Priority 2: Environment variable
    env_key = os.getenv('OPENAI_API_KEY')
    if env_key:
        return env_key
    
    # Priority 3: Local config file
    local_key = load_api_key()
    if local_key:
        return local_key
    
    # Priority 4: Session state (temporary)
    if 'api_key' in st.session_state and st.session_state.api_key:
        return st.session_state.api_key
    
    return None


def generate_pdf(requirements: dict) -> bytes:
    """Generate PDF from requirements dictionary."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=8,
        spaceBefore=12
    )
    
    # Title
    story.append(Paragraph("Requirements Extracted from Meeting", title_style))
    story.append(Paragraph(f"<i>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</i>", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Functional Requirements
    if requirements.get('functional_requirements'):
        story.append(Paragraph("Functional Requirements", heading_style))
        for req in requirements['functional_requirements']:
            story.append(Paragraph(f"<b>{req.get('id', 'N/A')}</b>", styles['Heading3']))
            story.append(Paragraph(f"<b>Description:</b> {req.get('description', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Priority:</b> {req.get('priority', 'Not specified')}", styles['Normal']))
            story.append(Paragraph(f"<b>Source:</b> {req.get('speaker', 'Unknown')}", styles['Normal']))
            if req.get('context'):
                story.append(Paragraph(f"<b>Context:</b> {req.get('context')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
    
    # Non-Functional Requirements
    if requirements.get('non_functional_requirements'):
        story.append(PageBreak())
        story.append(Paragraph("Non-Functional Requirements", heading_style))
        for req in requirements['non_functional_requirements']:
            story.append(Paragraph(f"<b>{req.get('id', 'N/A')}</b>", styles['Heading3']))
            story.append(Paragraph(f"<b>Description:</b> {req.get('description', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Priority:</b> {req.get('priority', 'Not specified')}", styles['Normal']))
            story.append(Paragraph(f"<b>Source:</b> {req.get('speaker', 'Unknown')}", styles['Normal']))
            if req.get('context'):
                story.append(Paragraph(f"<b>Context:</b> {req.get('context')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
    
    # Business Rules
    if requirements.get('business_rules'):
        story.append(PageBreak())
        story.append(Paragraph("Business Rules", heading_style))
        for rule in requirements['business_rules']:
            story.append(Paragraph(f"<b>{rule.get('id', 'N/A')}</b>", styles['Heading3']))
            story.append(Paragraph(f"<b>Rule:</b> {rule.get('description', rule.get('rule', 'N/A'))}", styles['Normal']))
            story.append(Paragraph(f"<b>Source:</b> {rule.get('speaker', 'Unknown')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
    
    # Action Items
    if requirements.get('action_items'):
        story.append(PageBreak())
        story.append(Paragraph("Action Items", heading_style))
        data = [['ID', 'Task', 'Owner', 'Deadline', 'Status']]
        for item in requirements['action_items']:
            data.append([
                item.get('id', 'N/A'),
                item.get('task', 'N/A'),
                item.get('owner', 'TBD'),
                item.get('deadline', 'TBD'),
                item.get('status', 'Open')
            ])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(table)
        story.append(Spacer(1, 0.2*inch))
    
    # Decisions
    if requirements.get('decisions'):
        story.append(PageBreak())
        story.append(Paragraph("Decisions", heading_style))
        for decision in requirements['decisions']:
            story.append(Paragraph(f"<b>{decision.get('id', 'N/A')}</b>", styles['Heading3']))
            story.append(Paragraph(f"<b>Decision:</b> {decision.get('decision', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Rationale:</b> {decision.get('rationale', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Decision Maker:</b> {decision.get('decision_maker', 'Unknown')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
    
    # Stakeholders
    if requirements.get('stakeholders'):
        story.append(PageBreak())
        story.append(Paragraph("Stakeholders", heading_style))
        for stakeholder in requirements['stakeholders']:
            story.append(Paragraph(f"<b>{stakeholder.get('name', 'Unknown')}</b>", styles['Heading3']))
            story.append(Paragraph(f"<b>Role:</b> {stakeholder.get('role', 'N/A')}", styles['Normal']))
            story.append(Paragraph(f"<b>Interests:</b> {stakeholder.get('interests', 'N/A')}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel(requirements: dict) -> bytes:
    """Generate Excel file from requirements dictionary."""
    if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
        raise ImportError("pandas and openpyxl are required for Excel export. Install them with: pip install pandas openpyxl")
    
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Functional Requirements
        if requirements.get('functional_requirements'):
            df_fr = pd.DataFrame(requirements['functional_requirements'])
            df_fr.to_excel(writer, sheet_name='Functional Requirements', index=False)
            worksheet = writer.sheets['Functional Requirements']
            # Format header
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Non-Functional Requirements
        if requirements.get('non_functional_requirements'):
            df_nfr = pd.DataFrame(requirements['non_functional_requirements'])
            df_nfr.to_excel(writer, sheet_name='Non-Functional Requirements', index=False)
            worksheet = writer.sheets['Non-Functional Requirements']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Business Rules
        if requirements.get('business_rules'):
            df_br = pd.DataFrame(requirements['business_rules'])
            df_br.to_excel(writer, sheet_name='Business Rules', index=False)
            worksheet = writer.sheets['Business Rules']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Action Items
        if requirements.get('action_items'):
            df_ai = pd.DataFrame(requirements['action_items'])
            df_ai.to_excel(writer, sheet_name='Action Items', index=False)
            worksheet = writer.sheets['Action Items']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Decisions
        if requirements.get('decisions'):
            df_dec = pd.DataFrame(requirements['decisions'])
            df_dec.to_excel(writer, sheet_name='Decisions', index=False)
            worksheet = writer.sheets['Decisions']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Stakeholders
        if requirements.get('stakeholders'):
            df_stk = pd.DataFrame(requirements['stakeholders'])
            df_stk.to_excel(writer, sheet_name='Stakeholders', index=False)
            worksheet = writer.sheets['Stakeholders']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    buffer.seek(0)
    return buffer.getvalue()

def initialize_session_state():
    """Initialize session state variables."""
    if 'requirements' not in st.session_state:
        st.session_state.requirements = None
    if 'transcript_text' not in st.session_state:
        st.session_state.transcript_text = None
    if 'messages_parsed' not in st.session_state:
        st.session_state.messages_parsed = None
    if 'partial_requirements' not in st.session_state:
        st.session_state.partial_requirements = []
    if 'processing_status' not in st.session_state:
        st.session_state.processing_status = None
    
    # Load API key from all sources on startup
    if 'api_key_loaded' not in st.session_state:
        auto_key = get_api_key_from_all_sources()
        if auto_key:
            st.session_state.api_key = auto_key
            st.session_state.api_key_source = "auto_loaded"
        st.session_state.api_key_loaded = True


def parse_transcript_file(uploaded_file):
    """Parse uploaded transcript file."""
    parser = TranscriptParser()
    
    # Save uploaded file temporarily
    file_extension = Path(uploaded_file.name).suffix.lower()
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        tmp_path = tmp_file.name
    
    try:
        # Parse based on file extension
        if file_extension == '.vtt':
            messages = parser.parse_vtt(tmp_path)
        elif file_extension == '.json':
            messages = parser.parse_json(tmp_path)
        else:
            messages = parser.parse_text(tmp_path)
        
        return messages, None
    except Exception as e:
        return None, str(e)
    finally:
        # Clean up temp file
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def parse_transcript_text(text):
    """Parse transcript from text input."""
    parser = TranscriptParser()
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.txt', encoding='utf-8') as tmp_file:
        tmp_file.write(text)
        tmp_path = tmp_file.name
    
    try:
        messages = parser.parse_text(tmp_path)
        return messages, None
    except Exception as e:
        return None, str(e)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def extract_audio_from_video(video_path: str, output_audio_path: str = None) -> str:
    """Extract audio from video file using MoviePy or FFmpeg directly."""
    if output_audio_path is None:
        output_audio_path = video_path.rsplit('.', 1)[0] + '.mp3'
    
    # Get FFmpeg path
    ffmpeg_path = None
    try:
        import imageio_ffmpeg
        ffmpeg_path = imageio_ffmpeg.get_ffmpeg_exe()
    except:
        # Try system FFmpeg
        import shutil
        ffmpeg_path = shutil.which('ffmpeg')
    
    if not ffmpeg_path:
        raise Exception("FFmpeg not found. Please install FFmpeg or imageio-ffmpeg.")
    
    # First, try to validate the video file using FFmpeg directly
    try:
        import subprocess
        # Check if video file is valid and readable
        check_cmd = [
            ffmpeg_path,
            '-i', video_path,
            '-t', '0.1',  # Just probe first 0.1 seconds
            '-f', 'null',
            '-'
        ]
        result = subprocess.run(
            check_cmd,
            capture_output=True,
            timeout=10
        )
        if result.returncode != 0:
            error_msg = result.stderr.decode('utf-8', errors='ignore')
            if 'Invalid data found' in error_msg or 'could not find codec parameters' in error_msg:
                raise Exception(f"Video file appears to be corrupted or in an unsupported format. FFmpeg error: {error_msg[:200]}")
    except subprocess.TimeoutExpired:
        raise Exception("Video file validation timed out. The file may be corrupted or too large.")
    except Exception as e:
        if "corrupted" in str(e).lower() or "unsupported format" in str(e).lower():
            raise
        # Continue if it's just a validation issue
    
    # Try MoviePy first (if available)
    if MOVIEPY_AVAILABLE:
        try:
            # Configure FFmpeg for MoviePy
            os.environ["FFMPEG_BINARY"] = ffmpeg_path
            os.environ["FFPROBE_BINARY"] = ffmpeg_path
            os.environ["FFPROBE"] = ffmpeg_path
            
            # Configure moviepy config directly
            try:
                import moviepy.config as mp_config
                mp_config.FFMPEG_BINARY = ffmpeg_path
                if hasattr(mp_config, 'FFPROBE_BINARY'):
                    mp_config.FFPROBE_BINARY = ffmpeg_path
            except:
                pass
            
            # Try to load video with MoviePy
            video = VideoFileClip(video_path, audio=True)
            if video.audio is None:
                video.close()
                raise Exception("Video file has no audio track.")
            
            audio = video.audio
            # moviepy 2.x doesn't support verbose parameter
            try:
                # Try without verbose (moviepy 2.x)
                audio.write_audiofile(output_audio_path, logger=None, codec='mp3')
            except TypeError:
                # Fallback for older versions
                try:
                    audio.write_audiofile(output_audio_path, verbose=False, logger=None, codec='mp3')
                except:
                    # Last resort: try with minimal parameters
                    audio.write_audiofile(output_audio_path, codec='mp3')
            audio.close()
            video.close()
            return output_audio_path
        except Exception as moviepy_error:
            error_msg = str(moviepy_error)
            # If MoviePy fails with frame reading error, fall back to FFmpeg directly
            if "failed to read the first frame" in error_msg.lower() or "corrupted" in error_msg.lower():
                # Fall through to FFmpeg direct method
                pass
            else:
                # For other MoviePy errors, try FFmpeg as fallback
                pass
    
    # Fallback: Use FFmpeg directly (more reliable for problematic videos)
    try:
        import subprocess
        ffmpeg_cmd = [
            ffmpeg_path,
            '-i', video_path,
            '-vn',  # No video
            '-acodec', 'libmp3lame',  # MP3 codec
            '-ab', '192k',  # Audio bitrate
            '-ar', '44100',  # Sample rate
            '-y',  # Overwrite output file
            output_audio_path
        ]
        
        result = subprocess.run(
            ffmpeg_cmd,
            capture_output=True,
            timeout=300  # 5 minute timeout
        )
        
        if result.returncode != 0:
            error_msg = result.stderr.decode('utf-8', errors='ignore')
            if 'Invalid data found' in error_msg or 'could not find codec parameters' in error_msg:
                raise Exception(f"Video file is corrupted or in an unsupported format. Details: {error_msg[:300]}")
            else:
                raise Exception(f"FFmpeg extraction failed: {error_msg[:300]}")
        
        # Verify output file was created
        if not os.path.exists(output_audio_path) or os.path.getsize(output_audio_path) == 0:
            raise Exception("Audio extraction completed but output file is empty or missing.")
        
        return output_audio_path
    except subprocess.TimeoutExpired:
        raise Exception("Audio extraction timed out. The video file may be too large or corrupted.")
    except Exception as e:
        raise Exception(f"Error extracting audio from video: {str(e)}")


def transcribe_audio_local_whisper(audio_path: str, progress_callback=None, model_size="base"):
    """Transcribe audio using local Whisper (no API key needed)."""
    if not WHISPER_LOCAL_AVAILABLE:
        raise ImportError("Local Whisper is not available. Install it with: pip install openai-whisper")
    
    try:
        # Check for NumPy before proceeding
        try:
            import numpy
        except ImportError:
            raise ImportError("NumPy is required for Whisper. Install it with: pip install numpy")
        
        # Configure ffmpeg path for Whisper (use imageio_ffmpeg's bundled ffmpeg)
        try:
            import imageio_ffmpeg
            ffmpeg_path = imageio_ffmpeg.get_ffmpeg_exe()
            
            # Set environment variables for ffmpeg
            os.environ["FFMPEG_BINARY"] = ffmpeg_path
            os.environ["FFMPEG_PATH"] = ffmpeg_path
            
            # Add ffmpeg directory to PATH so subprocess calls can find it
            ffmpeg_dir = os.path.dirname(ffmpeg_path)
            current_path = os.environ.get("PATH", "")
            if ffmpeg_dir not in current_path:
                os.environ["PATH"] = ffmpeg_dir + os.pathsep + current_path
            
            # Create a symlink named 'ffmpeg' in a temp directory that Whisper can find
            import os.path as osp
            temp_bin_dir = osp.join(osp.expanduser("~"), ".requirements_extractor_bin")
            os.makedirs(temp_bin_dir, exist_ok=True)
            ffmpeg_link = osp.join(temp_bin_dir, "ffmpeg")
            
            # Create symlink if it doesn't exist or is broken
            if not osp.exists(ffmpeg_link) or not osp.islink(ffmpeg_link):
                try:
                    if osp.exists(ffmpeg_link):
                        os.remove(ffmpeg_link)
                    os.symlink(ffmpeg_path, ffmpeg_link)
                except Exception:
                    # If symlink fails, try creating a wrapper script
                    with open(ffmpeg_link, 'w') as f:
                        f.write(f"""#!/bin/bash
exec "{ffmpeg_path}" "$@"
""")
                    os.chmod(ffmpeg_link, 0o755)
            
            # Add temp bin directory to PATH (at the beginning for priority)
            if temp_bin_dir not in os.environ.get("PATH", ""):
                os.environ["PATH"] = temp_bin_dir + os.pathsep + os.environ.get("PATH", "")
                
        except Exception as ffmpeg_error:
            # If we can't set ffmpeg path, continue - might work with system ffmpeg
            pass
        
        if progress_callback:
            progress_callback(0.1, f"📥 Loading Whisper model ({model_size})...")
        
        # Load Whisper model
        model = whisper.load_model(model_size)
        
        if progress_callback:
            progress_callback(0.2, "🎤 Starting transcription (this may take several minutes for large files)...")
        
        # Get audio duration for better progress estimation
        try:
            import subprocess
            import json
            ffprobe_cmd = [
                ffmpeg_path if 'ffmpeg_path' in locals() else 'ffprobe',
                '-v', 'quiet',
                '-print_format', 'json',
                '-show_format',
                audio_path
            ]
            probe_result = subprocess.run(ffprobe_cmd, capture_output=True, text=True, timeout=10)
            if probe_result.returncode == 0:
                audio_info = json.loads(probe_result.stdout)
                duration = float(audio_info.get('format', {}).get('duration', 0))
                if duration > 0:
                    estimated_minutes = duration / 60
                    if progress_callback:
                        progress_callback(0.25, f"🎤 Transcribing {estimated_minutes:.1f} minutes of audio (this may take {estimated_minutes * 2:.0f}-{estimated_minutes * 5:.0f} minutes)...")
        except:
            if progress_callback:
                progress_callback(0.25, "🎤 Transcribing audio (this may take several minutes)...")
        
        # Note: Removed background thread updates - Streamlit doesn't support UI updates from threads
        # Progress will be shown at key checkpoints instead
        if progress_callback:
            progress_callback(0.3, "🎤 Transcribing audio... This may take several minutes for large files...")
        
        # Transcribe audio - explicitly set ffmpeg path if available
        # Use verbose=False to reduce output, and fp16=False for better compatibility
        try:
            try:
                result = model.transcribe(
                    audio_path,
                    verbose=False,
                    fp16=False,  # Use fp32 for better compatibility
                    language=None,  # Auto-detect language
                    task="transcribe"
                )
            except FileNotFoundError as e:
                if "ffmpeg" in str(e).lower():
                    # Try to use imageio_ffmpeg's ffmpeg
                    try:
                        import imageio_ffmpeg
                        ffmpeg_path = imageio_ffmpeg.get_ffmpeg_exe()
                        # Set as environment variable
                        os.environ["PATH"] = os.path.dirname(ffmpeg_path) + os.pathsep + os.environ.get("PATH", "")
                        os.environ["FFMPEG_BINARY"] = ffmpeg_path
                        # Retry transcription
                        result = model.transcribe(audio_path)
                    except Exception as retry_error:
                        raise Exception(
                            f"FFmpeg not found. Whisper needs FFmpeg to process audio.\n\n"
                            f"Error: {str(e)}\n\n"
                            f"Solution: Install FFmpeg:\n"
                            f"  macOS: brew install ffmpeg\n"
                            f"  Linux: sudo apt-get install ffmpeg\n"
                            f"  Or ensure imageio_ffmpeg is installed: pip install imageio-ffmpeg"
                        )
                else:
                    raise
        finally:
            # Cleanup completed
            pass
        
        if progress_callback:
            progress_callback(0.85, "📝 Processing transcription results...")
        
        # Extract text from result
        text = result.get("text", "")
        
        if progress_callback:
            progress_callback(0.9, f"✅ Transcription complete! ({len(text)} characters)")
        
        if not text or len(text.strip()) == 0:
            raise Exception("Received empty transcript from Whisper")
        
        return text, []
        
    except ImportError as e:
        # Re-raise import errors with helpful message
        raise Exception(f"Missing dependency: {str(e)}\n\nPlease install: pip install numpy")
    except Exception as e:
        error_msg = str(e)
        # Check if it's a NumPy-related error
        if "numpy" in error_msg.lower() or "NumPy" in error_msg:
            raise Exception(
                f"NumPy error: {error_msg}\n\n"
                "This might be a NumPy version compatibility issue.\n"
                "Try: pip install 'numpy<2.0'"
            )
        # Check if it's an ffmpeg error
        if "ffmpeg" in error_msg.lower() or "No such file" in error_msg:
            raise Exception(
                f"FFmpeg error: {error_msg}\n\n"
                "Whisper needs FFmpeg to process audio files.\n\n"
                "Install FFmpeg:\n"
                "  macOS: brew install ffmpeg\n"
                "  Linux: sudo apt-get install ffmpeg\n"
                "  Windows: Download from https://ffmpeg.org\n\n"
                "Or install imageio-ffmpeg: pip install imageio-ffmpeg"
            )
        raise Exception(f"Error transcribing with local Whisper: {error_msg}")


def transcribe_audio_with_whisper(audio_path: str, api_key: str = None, use_local: bool = False, progress_callback=None):
    """Transcribe audio file using OpenAI Whisper API. Handles large files by chunking."""
    try:
        from openai import OpenAI
        import json
        
        # Validate API key
        if not api_key or not api_key.strip():
            raise Exception("❌ OpenAI API key is missing. Please provide your API key in the sidebar.")
        
        # Validate API key format (should start with sk-)
        if not api_key.startswith('sk-'):
            raise Exception("❌ Invalid API key format. OpenAI API keys should start with 'sk-'. Please check your API key.")
        
        # Test API key with a simple request first
        try:
            client = OpenAI(api_key=api_key)
            # Try to list models to verify API key works
            # This is a lightweight test that doesn't cost anything
            try:
                models = client.models.list(limit=1)
                # If we get here, the API key is valid
                if progress_callback:
                    progress_callback(0.05, "✅ API key validated")
            except Exception as model_test_error:
                error_str = str(model_test_error)
                # If models.list fails, that's okay - we'll still try transcription
                # Some API keys might not have model list access
                if progress_callback:
                    progress_callback(0.05, "⚠️ Could not validate API key (will still try transcription)")
        except Exception as key_test_error:
            error_str = str(key_test_error)
            if "Invalid OpenAI API key" in error_str:
                raise  # Re-raise our custom error
            # Otherwise, continue - might be a network issue
        
        # Check file size (Whisper API limit is 25MB)
        file_size_mb = os.path.getsize(audio_path) / (1024 * 1024)
        max_size_mb = 25  # OpenAI Whisper API limit
        
        if file_size_mb <= max_size_mb:
            # File is small enough, transcribe directly
            with open(audio_path, 'rb') as audio_file:
                try:
                    # Verify file is readable
                    file_size = os.path.getsize(audio_path)
                    if file_size == 0:
                        raise Exception("Audio file is empty or corrupted")
                    
                    # Try transcription - catch all errors including JSON parsing
                    try:
                        # Reset file pointer to beginning
                        audio_file.seek(0)
                        
                        # Make the API call with explicit parameters
                        transcript = client.audio.transcriptions.create(
                            model="whisper-1",
                            file=audio_file,
                            response_format="text"  # Explicitly request text format
                        )
                    except Exception as create_error:
                        error_str = str(create_error)
                        error_type = type(create_error).__name__
                        
                        # Try to extract detailed error information
                        debug_info = []
                        debug_info.append(f"Error Type: {error_type}")
                        debug_info.append(f"Error Message: {error_str}")
                        
                        # Check for OpenAI API error attributes
                        if hasattr(create_error, 'response'):
                            try:
                                response = create_error.response
                                if hasattr(response, 'status_code'):
                                    debug_info.append(f"HTTP Status: {response.status_code}")
                                if hasattr(response, 'text'):
                                    response_text = response.text[:1000]
                                    debug_info.append(f"Response Text: {response_text}")
                                if hasattr(response, 'headers'):
                                    debug_info.append(f"Response Headers: {dict(response.headers)}")
                            except Exception as e:
                                debug_info.append(f"Could not extract response details: {e}")
                        
                        # Check for body attribute
                        if hasattr(create_error, 'body'):
                            try:
                                body = create_error.body
                                if isinstance(body, dict):
                                    debug_info.append(f"Error Body: {body}")
                                elif isinstance(body, str):
                                    debug_info.append(f"Error Body: {body[:1000]}")
                            except:
                                pass
                        
                        # Check if it's a JSON-related error (most common cause: invalid API key)
                        if "Expecting value" in error_str or "JSON" in error_str or "JSONDecodeError" in error_type or "json" in error_type.lower() or "decode" in error_str.lower():
                            # This error almost always means invalid API key
                            debug_details = "\n".join(debug_info)
                            raise Exception(
                                "❌ INVALID API KEY DETECTED\n\n"
                                "The error 'Expecting value: line 1 column 1 (char 0)' means the API returned\n"
                                "an empty or non-JSON response, which typically indicates an INVALID API KEY.\n\n"
                                "🔧 FIX THIS:\n"
                                "1. Open https://platform.openai.com/api-keys in your browser\n"
                                "2. Click 'Create new secret key'\n"
                                "3. Copy the NEW key (starts with 'sk-')\n"
                                "4. Paste it in the sidebar of this app\n"
                                "5. Make sure billing is set up: https://platform.openai.com/account/billing\n\n"
                                f"Debug Information:\n{debug_details}\n\n"
                                f"Original error: {error_str}"
                            )
                        raise
                    
                    # Handle response - OpenAI SDK returns an object with .text attribute
                    if hasattr(transcript, 'text'):
                        text = transcript.text
                    elif isinstance(transcript, str):
                        text = transcript
                    else:
                        text = str(transcript)
                    
                    if not text or len(text.strip()) == 0:
                        raise Exception("Received empty transcript from Whisper API")
                    
                    return text, []
                    
                except Exception as api_error:
                    error_msg = str(api_error)
                    error_type = type(api_error).__name__
                    
                    # Get more details from the exception if available
                    error_details = ""
                    raw_response = ""
                    
                    # Try to extract response details from OpenAI API error
                    if hasattr(api_error, 'response'):
                        try:
                            response = api_error.response
                            if hasattr(response, 'text'):
                                raw_response = response.text[:500]
                            elif hasattr(response, 'content'):
                                raw_response = str(response.content)[:500]
                            elif hasattr(response, 'body'):
                                raw_response = str(response.body)[:500]
                            
                            if hasattr(response, 'status_code'):
                                error_details += f"\nHTTP Status: {response.status_code}"
                            if hasattr(response, 'headers'):
                                error_details += f"\nResponse Headers: {str(response.headers)[:200]}"
                        except Exception as e:
                            pass
                    
                    # Also check for body attribute directly on the error
                    if hasattr(api_error, 'body'):
                        try:
                            if isinstance(api_error.body, dict):
                                raw_response = str(api_error.body)[:500]
                            elif isinstance(api_error.body, str):
                                raw_response = api_error.body[:500]
                        except:
                            pass
                    
                    # Check for common API errors
                    if "Invalid API key" in error_msg or "401" in error_msg or "authentication" in error_msg.lower() or "incorrect API key" in error_msg.lower() or "invalid_api_key" in error_msg.lower() or "401" in str(raw_response):
                        raise Exception("❌ Invalid OpenAI API key. Please check your API key in the sidebar and ensure it's correct.\n\nGet your API key at: https://platform.openai.com/api-keys")
                    elif "insufficient_quota" in error_msg or "429" in error_msg or "quota" in error_msg.lower() or "billing" in error_msg.lower() or "rate_limit" in error_msg.lower() or "429" in str(raw_response):
                        raise Exception("❌ OpenAI API quota exceeded or rate limit reached. Please check your account billing at https://platform.openai.com/account/billing")
                    elif "file_size" in error_msg.lower() or "too large" in error_msg.lower():
                        raise Exception(f"Audio file is too large ({file_size_mb:.2f}MB). The file will be chunked automatically.")
                    elif "JSON" in error_type or "json" in error_msg.lower() or "Expecting value" in error_msg or "decode" in error_msg.lower() or "JSONDecodeError" in error_type:
                        # JSON parsing error - likely API returned error HTML or empty response
                        detailed_msg = f"❌ API communication error: {error_msg}"
                        detailed_msg += "\n\nThis usually means:"
                        detailed_msg += "\n1. Invalid or expired API key (most common)"
                        detailed_msg += "\n2. Network/proxy/firewall blocking the request"
                        detailed_msg += "\n3. OpenAI API service issue"
                        detailed_msg += "\n4. API key doesn't have access to Whisper API"
                        detailed_msg += f"\n\nPlease verify your API key at: https://platform.openai.com/api-keys"
                        if error_details:
                            detailed_msg += error_details
                        if raw_response:
                            detailed_msg += f"\n\nRaw API Response (first 500 chars):\n{raw_response}"
                        raise Exception(detailed_msg)
                    elif "timeout" in error_msg.lower():
                        raise Exception("❌ Request timed out. The audio file might be too large. Try a smaller file or check your network connection.")
                    elif "connection" in error_msg.lower() or "network" in error_msg.lower():
                        raise Exception("❌ Network connection error. Please check your internet connection and try again.")
                    else:
                        detailed_msg = f"❌ Whisper API error: {error_msg}"
                        detailed_msg += "\n\nPlease check:"
                        detailed_msg += "\n1. Your API key is valid and active"
                        detailed_msg += "\n2. You have sufficient credits"
                        detailed_msg += "\n3. Your network connection is stable"
                        detailed_msg += "\n4. The audio file format is supported"
                        if error_details:
                            detailed_msg += error_details
                        if raw_response:
                            detailed_msg += f"\n\nRaw API Response (first 500 chars):\n{raw_response}"
                        raise Exception(detailed_msg)
        else:
            # File is too large, need to compress or chunk it
            if not PYDUB_AVAILABLE:
                # Try to compress using ffmpeg first
                if progress_callback:
                    progress_callback(0.2, "🔄 Compressing audio file (too large for direct upload)...")
                
                try:
                    import subprocess
                    compressed_path = audio_path.rsplit('.', 1)[0] + '_compressed.mp3'
                    
                    # Compress audio to lower bitrate (64kbps mono should be under 25MB for most files)
                    cmd = [
                        'ffmpeg', '-y', '-i', audio_path,
                        '-ac', '1',  # Mono
                        '-ar', '16000',  # 16kHz sample rate (good for speech)
                        '-b:a', '64k',  # 64kbps bitrate
                        compressed_path
                    ]
                    
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
                    
                    if result.returncode == 0 and os.path.exists(compressed_path):
                        compressed_size_mb = os.path.getsize(compressed_path) / (1024 * 1024)
                        if progress_callback:
                            progress_callback(0.3, f"✅ Compressed to {compressed_size_mb:.1f}MB. Transcribing...")
                        
                        with open(compressed_path, 'rb') as audio_file:
                            transcript = client.audio.transcriptions.create(
                                model="whisper-1",
                                file=audio_file,
                                response_format="verbose_json"
                            )
                        
                        # Clean up compressed file
                        try:
                            os.remove(compressed_path)
                        except:
                            pass
                        
                        return transcript.text, [{"start": 0, "end": 0, "text": transcript.text}]
                    else:
                        raise Exception(f"FFmpeg compression failed: {result.stderr}")
                        
                except subprocess.TimeoutExpired:
                    raise Exception("Audio compression timed out. File may be too large.")
                except FileNotFoundError:
                    raise Exception("FFmpeg not found. Cannot process large audio files.")
                except Exception as e:
                    raise Exception(f"Error compressing audio: {str(e)}")
            
            if progress_callback:
                progress_callback(0.1, "📦 Splitting large audio file into chunks...")
            
            # Load audio file
            audio = AudioSegment.from_file(audio_path)
            duration_seconds = len(audio) / 1000.0
            
            # Calculate chunk size (aim for ~20MB chunks to be safe)
            # Approximate: 1 minute of audio ≈ 1MB (compressed MP3)
            # So 20 minutes ≈ 20MB
            chunk_duration_ms = 20 * 60 * 1000  # 20 minutes in milliseconds
            total_chunks = int(duration_seconds / (chunk_duration_ms / 1000)) + 1
            
            all_text_parts = []
            all_segments = []
            
            # Process in chunks
            for i in range(total_chunks):
                start_ms = i * chunk_duration_ms
                end_ms = min((i + 1) * chunk_duration_ms, len(audio))
                
                if start_ms >= len(audio):
                    break
                
                if progress_callback:
                    progress = 0.1 + (i / total_chunks) * 0.8
                    progress_callback(progress, f"🎤 Transcribing chunk {i+1}/{total_chunks}...")
                
                # Extract chunk
                chunk = audio[start_ms:end_ms]
                
                # Save chunk temporarily
                chunk_path = audio_path.rsplit('.', 1)[0] + f'_chunk_{i}.mp3'
                chunk.export(chunk_path, format="mp3")
                
                try:
                    # Transcribe chunk
                    with open(chunk_path, 'rb') as chunk_file:
                        try:
                            transcript = client.audio.transcriptions.create(
                                model="whisper-1",
                                file=chunk_file
                            )
                            chunk_text = transcript.text if hasattr(transcript, 'text') else (transcript if isinstance(transcript, str) else str(transcript))
                            chunk_segments = []
                            
                            if not chunk_text or len(chunk_text.strip()) == 0:
                                # Skip empty chunks
                                continue
                        except Exception as chunk_api_error:
                            error_msg = str(chunk_api_error)
                            error_type = type(chunk_api_error).__name__
                            
                            # Check if it's a JSON-related error (invalid API key)
                            if "JSON" in error_msg or "Expecting value" in error_msg or "decode" in error_msg.lower() or "JSONDecodeError" in error_type:
                                # Critical error - invalid API key, stop processing
                                raise Exception(
                                    "❌ INVALID API KEY DETECTED\n\n"
                                    "The error 'Expecting value: line 1 column 1 (char 0)' means the API returned\n"
                                    "an empty or non-JSON response, which typically indicates an INVALID API KEY.\n\n"
                                    "🔧 FIX THIS:\n"
                                    "1. Open https://platform.openai.com/api-keys in your browser\n"
                                    "2. Click 'Create new secret key'\n"
                                    "3. Copy the NEW key (starts with 'sk-')\n"
                                    "4. Paste it in the sidebar of this app\n"
                                    "5. Make sure billing is set up: https://platform.openai.com/account/billing\n\n"
                                    f"Original error: {error_msg}"
                                )
                            # For other errors, skip this chunk but continue
                            if progress_callback:
                                progress_callback(progress, f"⚠️ Warning: Skipped chunk {i+1} due to error: {error_msg}")
                            continue
                    
                    # Adjust segment timestamps (segments are dicts from API)
                    offset_seconds = start_ms / 1000.0
                    for segment in chunk_segments:
                        if isinstance(segment, dict):
                            if 'start' in segment:
                                segment['start'] += offset_seconds
                            if 'end' in segment:
                                segment['end'] += offset_seconds
                        else:
                            # If it's an object, try to update attributes
                            if hasattr(segment, 'start'):
                                segment.start += offset_seconds
                            if hasattr(segment, 'end'):
                                segment.end += offset_seconds
                    
                    all_text_parts.append(chunk_text)
                    all_segments.extend(chunk_segments)
                    
                except Exception as chunk_error:
                    # Re-raise if it's our custom invalid API key error
                    error_msg = str(chunk_error)
                    if "INVALID API KEY" in error_msg:
                        raise
                    # If chunk fails for other reasons, log and continue with other chunks
                    if "JSON" in error_msg or "Expecting value" in error_msg or "decode" in error_msg.lower():
                        # Critical error - likely API key issue, stop processing
                        raise Exception(f"❌ API error while processing chunk {i+1}: {error_msg}\n\nThis usually indicates an invalid API key. Please check your API key.")
                    # For other errors, skip this chunk but continue
                    if progress_callback:
                        progress_callback(progress, f"⚠️ Warning: Skipped chunk {i+1} due to error")
                    continue
                finally:
                    # Clean up chunk file
                    if os.path.exists(chunk_path):
                        os.unlink(chunk_path)
            
            # Combine all transcripts
            combined_text = '\n'.join(all_text_parts)
            
            if progress_callback:
                progress_callback(0.95, "✅ Combining transcripts...")
            
            return combined_text, all_segments
            
    except Exception as e:
        error_msg = str(e)
        # Don't double-wrap our custom error messages - preserve them completely
        if "INVALID API KEY" in error_msg or "FIX THIS:" in error_msg or "Debug Information:" in error_msg:
            raise  # Re-raise our custom error as-is (it already has all the details)
        # For JSON errors, provide detailed help
        if "Expecting value" in error_msg or "JSON" in error_msg or "JSONDecodeError" in str(type(e).__name__):
            raise Exception(
                f"Error transcribing audio: {error_msg}\n\n"
                "This JSON parsing error typically means:\n"
                "1. ❌ INVALID API KEY (most common)\n"
                "2. Network/proxy blocking the request\n"
                "3. API service issue\n\n"
                "Please:\n"
                "1. Verify your API key at https://platform.openai.com/api-keys\n"
                "2. Generate a NEW key if needed\n"
                "3. Check billing at https://platform.openai.com/account/billing"
            )
        # Otherwise, wrap it but preserve the full message
        raise Exception(f"Error transcribing audio: {error_msg}")


def process_audio_file(uploaded_file, api_key: str = None, progress_bar=None, status_text=None, use_local: bool = False):
    """Process audio file directly: transcribe. Handles files of any size."""
    file_extension = Path(uploaded_file.name).suffix.lower()
    
    # Save audio temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_audio:
        tmp_audio.write(uploaded_file.getvalue())
        tmp_audio_path = tmp_audio.name
    
    transcript_text = None
    
    def update_progress(progress, message):
        """Helper to update progress bar and status."""
        if progress_bar:
            progress_bar.progress(progress)
        if status_text:
            status_text.info(message)
    
    try:
        # Step 1: Get file info
        audio_size_mb = os.path.getsize(tmp_audio_path) / (1024 * 1024)
        update_progress(0.1, f"🎤 Processing audio file ({audio_size_mb:.2f} MB)...")
        
        # Step 2: Transcribe audio (with chunking for large files)
        def progress_callback(progress, message):
            # Map 0.1-0.95 to 0.1-0.9 range
            mapped_progress = 0.1 + (progress * 0.8)
            update_progress(mapped_progress, message)
        
        # Determine model size based on audio file size
        if audio_size_mb > 50:  # Large file
            model_size = "tiny"  # Fastest model for large files
            update_progress(0.15, f"📊 Large file detected ({audio_size_mb:.1f} MB). Using fast 'tiny' model for quicker transcription...")
        elif audio_size_mb > 20:  # Medium file
            model_size = "base"  # Balanced model
            update_progress(0.15, f"📊 Medium file ({audio_size_mb:.1f} MB). Using 'base' model...")
        else:  # Small file
            model_size = "base"  # Can use base or small
            update_progress(0.15, f"📊 Small file ({audio_size_mb:.1f} MB). Using 'base' model...")
        
        # Use local Whisper if available and requested, otherwise use API
        if use_local and WHISPER_LOCAL_AVAILABLE:
            transcript_text, segments = transcribe_audio_local_whisper(
                tmp_audio_path,
                progress_callback=progress_callback,
                model_size=model_size
            )
        else:
            if not api_key:
                raise Exception("❌ API key is required when using OpenAI API. Please provide your API key in the sidebar or enable local Whisper transcription.")
            transcript_text, segments = transcribe_audio_with_whisper(
                tmp_audio_path, 
                api_key,
                use_local=False,
                progress_callback=progress_callback
            )
        
        update_progress(0.9, "📝 Processing transcript...")
        
        # Step 3: Parse transcript
        parser = TranscriptParser()
        # Create a simple transcript format from Whisper output
        # Whisper doesn't identify speakers, so we'll create a single speaker transcript
        lines = transcript_text.split('\n')
        messages = []
        for line in lines:
            line = line.strip()
            if line:
                messages.append({
                    'speaker': 'Speaker',
                    'text': line,
                    'timestamp': None
                })
        
        update_progress(1.0, "✅ Transcription complete!")
        
        return messages, None
        
    except Exception as e:
        return None, str(e)
    finally:
        # Clean up temp files
        if tmp_audio_path and os.path.exists(tmp_audio_path):
            try:
                os.unlink(tmp_audio_path)
            except:
                pass


def process_video_file(uploaded_file, api_key: str = None, progress_bar=None, status_text=None, use_local: bool = False):
    """Process video file: extract audio and transcribe. Handles files of any size."""
    file_extension = Path(uploaded_file.name).suffix.lower()
    
    # Save video temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_video:
        tmp_video.write(uploaded_file.getvalue())
        tmp_video_path = tmp_video.name
    
    tmp_audio_path = None
    transcript_text = None
    
    def update_progress(progress, message):
        """Helper to update progress bar and status."""
        if progress_bar:
            progress_bar.progress(progress)
        if status_text:
            status_text.info(message)
    
    try:
        # Step 1: Extract audio
        update_progress(0.1, "🎬 Extracting audio from video...")
        
        tmp_audio_path = tmp_video_path.rsplit('.', 1)[0] + '.mp3'
        extract_audio_from_video(tmp_video_path, tmp_audio_path)
        
        audio_size_mb = os.path.getsize(tmp_audio_path) / (1024 * 1024)
        update_progress(0.2, f"🎤 Audio extracted ({audio_size_mb:.2f} MB). Starting transcription...")
        
        # Step 2: Transcribe audio (with chunking for large files)
        def progress_callback(progress, message):
            # Map 0.1-0.95 to 0.2-0.9 range
            mapped_progress = 0.2 + (progress * 0.7)
            update_progress(mapped_progress, message)
        
        # Determine model size based on audio file size
        # For large files, use a smaller/faster model
        audio_size_mb = os.path.getsize(tmp_audio_path) / (1024 * 1024)
        if audio_size_mb > 50:  # Large file
            model_size = "tiny"  # Fastest model for large files
            update_progress(0.25, f"📊 Large file detected ({audio_size_mb:.1f} MB). Using fast 'tiny' model for quicker transcription...")
        elif audio_size_mb > 20:  # Medium file
            model_size = "base"  # Balanced model
            update_progress(0.25, f"📊 Medium file ({audio_size_mb:.1f} MB). Using 'base' model...")
        else:  # Small file
            model_size = "base"  # Can use base or small
            update_progress(0.25, f"📊 Small file ({audio_size_mb:.1f} MB). Using 'base' model...")
        
        # Use local Whisper if available and requested, otherwise use API
        if use_local and WHISPER_LOCAL_AVAILABLE:
            transcript_text, segments = transcribe_audio_local_whisper(
                tmp_audio_path,
                progress_callback=progress_callback,
                model_size=model_size
            )
        else:
            if not api_key:
                raise Exception("❌ API key is required when using OpenAI API. Please provide your API key in the sidebar or enable local Whisper transcription.")
            transcript_text, segments = transcribe_audio_with_whisper(
                tmp_audio_path, 
                api_key,
                use_local=False,
                progress_callback=progress_callback
            )
        
        update_progress(0.9, "📝 Processing transcript...")
        
        # Step 3: Parse transcript
        parser = TranscriptParser()
        # Create a simple transcript format from Whisper output
        # Whisper doesn't identify speakers, so we'll create a single speaker transcript
        lines = transcript_text.split('\n')
        messages = []
        for line in lines:
            line = line.strip()
            if line:
                messages.append({
                    'speaker': 'Speaker',
                    'text': line,
                    'timestamp': None
                })
        
        update_progress(1.0, "✅ Transcription complete!")
        
        return messages, None
        
    except Exception as e:
        return None, str(e)
    finally:
        # Clean up temp files
        for path in [tmp_video_path, tmp_audio_path]:
            if path and os.path.exists(path):
                try:
                    os.unlink(path)
                except:
                    pass


def extract_requirements(messages, api_key, model, use_ollama=False, ollama_model="llama3.2", chunk_size=50, progress_callback=None, feedback=None):
    """
    Extract requirements from parsed messages.
    Processes in chunks and generates incremental reports.
    
    Args:
        messages: List of message dictionaries
        api_key: API key for OpenAI (if not using Ollama)
        model: Model name for OpenAI
        use_ollama: Whether to use Ollama
        ollama_model: Ollama model name
        chunk_size: Number of messages to process at a time
        progress_callback: Function to call with (progress, message) updates
        feedback: Optional feedback or corrections to incorporate into extraction
    """
    try:
        extractor = RequirementsExtractor(
            api_key=api_key, 
            model=model,
            use_ollama=use_ollama,
            ollama_model=ollama_model
        )
        
        # If messages are small, process all at once
        if len(messages) <= chunk_size:
            if progress_callback:
                progress_callback(0.5, "🤖 Extracting requirements from transcript...")
            requirements = extractor.extract_requirements(messages, feedback=feedback)
            if progress_callback:
                progress_callback(1.0, "✅ Requirements extracted!")
            return requirements, None
        
        # For large transcripts, process in chunks
        total_chunks = (len(messages) + chunk_size - 1) // chunk_size
        all_requirements = {
            'functional_requirements': [],
            'non_functional_requirements': [],
            'business_rules': [],
            'action_items': [],
            'decisions': [],
            'stakeholders': []
        }
        
        for i in range(total_chunks):
            start_idx = i * chunk_size
            end_idx = min((i + 1) * chunk_size, len(messages))
            chunk_messages = messages[start_idx:end_idx]
            
            if progress_callback:
                progress = (i / total_chunks) * 0.9
                progress_callback(progress, f"🤖 Extracting requirements from chunk {i+1}/{total_chunks} ({len(chunk_messages)} messages)...")
            
            # Extract requirements for this chunk
            # Pass feedback to all chunks so guidance is applied consistently
            chunk_requirements = extractor.extract_requirements(chunk_messages, feedback=feedback)
            
            # Merge requirements
            for key in all_requirements.keys():
                if chunk_requirements.get(key):
                    all_requirements[key].extend(chunk_requirements[key])
            
            # Update session state with partial results
            if 'partial_requirements' in st.session_state:
                st.session_state.partial_requirements.append({
                    'chunk': i + 1,
                    'total_chunks': total_chunks,
                    'requirements': chunk_requirements,
                    'timestamp': datetime.now().isoformat()
                })
        
        if progress_callback:
            progress_callback(0.95, "📊 Combining all requirements...")
        
        # Deduplicate and merge requirements
        # Remove duplicates based on description
        seen_descriptions = set()
        for key in ['functional_requirements', 'non_functional_requirements', 'business_rules']:
            if all_requirements.get(key):
                unique_items = []
                for item in all_requirements[key]:
                    desc = item.get('description', item.get('rule', ''))
                    if desc and desc not in seen_descriptions:
                        seen_descriptions.add(desc)
                        unique_items.append(item)
                all_requirements[key] = unique_items
        
        # Merge stakeholders (by name)
        if all_requirements.get('stakeholders'):
            stakeholders_dict = {}
            for stakeholder in all_requirements['stakeholders']:
                name = stakeholder.get('name', 'Unknown')
                if name not in stakeholders_dict:
                    stakeholders_dict[name] = stakeholder
                else:
                    # Merge roles and interests
                    existing = stakeholders_dict[name]
                    if stakeholder.get('role') and not existing.get('role'):
                        existing['role'] = stakeholder['role']
                    if stakeholder.get('interests') and not existing.get('interests'):
                        existing['interests'] = stakeholder['interests']
            all_requirements['stakeholders'] = list(stakeholders_dict.values())
        
        if progress_callback:
            progress_callback(1.0, "✅ All requirements extracted and combined!")
        
        return all_requirements, None
        
    except Exception as e:
        return None, str(e)


def show_signup_page():
    """Display sign up page."""
    st.markdown("""
    <div style="text-align: center; padding: 2rem 0 1rem 0; background-color: #ffffff;">
        <div style="font-size: 5rem; margin-bottom: 0.5rem;">✨</div>
        <h1 style="
            font-size: 4.5rem;
            font-weight: 900;
            color: #000000 !important;
            margin: 0;
            letter-spacing: -0.03em;
            line-height: 1;
        ">ReqIQ</h1>
        <p style="
            font-size: 1.5rem;
            color: #333333 !important;
            margin: 0.75rem 0 0 0;
            font-weight: 500;
        ">Create Your Account</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("### 📝 Sign Up")
        
        email = st.text_input(
            "Email Address",
            placeholder="your.email@example.com",
            help="Enter your email address",
            key="signup_email"
        )
        
        password = st.text_input(
            "Password",
            type="password",
            placeholder="Enter a password (min 6 characters)",
            help="Password must be at least 6 characters",
            key="signup_password"
        )
        
        password_confirm = st.text_input(
            "Confirm Password",
            type="password",
            placeholder="Confirm your password",
            help="Re-enter your password",
            key="signup_password_confirm"
        )
        
        if st.button("🚀 Sign Up", type="primary", use_container_width=False, key="signup_button"):
            if not email or not password:
                st.error("❌ Please fill in all fields")
            elif password != password_confirm:
                st.error("❌ Passwords do not match")
            elif len(password) < 6:
                st.error("❌ Password must be at least 6 characters")
            else:
                if AUTH_AVAILABLE:
                    auth_manager = AuthManager()
                    result = auth_manager.register_user(email, password)
                    
                    if result.get("success"):
                        st.success("✅ Account created successfully!")
                        st.info("📧 Please sign in with your email and password")
                        # Switch to sign in
                        st.session_state.show_signin = True
                        st.rerun()
                    else:
                        st.error(f"❌ {result.get('error', 'Registration failed')}")
                else:
                    st.error("❌ Authentication system not available")
        
        st.markdown("---")
        st.markdown("### Already have an account?")
        if st.button("🔐 Sign In", use_container_width=False, key="goto_signin"):
            st.session_state.show_signin = True
            st.rerun()


def show_signin_page():
    """Display sign in page."""
    st.markdown("""
    <div style="text-align: center; padding: 2rem 0 1rem 0; background-color: #ffffff;">
        <div style="font-size: 5rem; margin-bottom: 0.5rem;">✨</div>
        <h1 style="
            font-size: 4.5rem;
            font-weight: 900;
            color: #000000 !important;
            margin: 0;
            letter-spacing: -0.03em;
            line-height: 1;
        ">ReqIQ</h1>
        <p style="
            font-size: 1.5rem;
            color: #333333 !important;
            margin: 0.75rem 0 0 0;
            font-weight: 500;
        ">Welcome Back</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("### 🔐 Sign In")
        
        email = st.text_input(
            "Email Address",
            placeholder="your.email@example.com",
            help="Enter your email address",
            key="signin_email"
        )
        
        password = st.text_input(
            "Password",
            type="password",
            placeholder="Enter your password",
            help="Enter your password",
            key="signin_password"
        )
        
        if st.button("🚀 Sign In", type="primary", use_container_width=False, key="signin_button"):
            if not email or not password:
                st.error("❌ Please enter email and password")
            else:
                if AUTH_AVAILABLE:
                    auth_manager = AuthManager()
                    result = auth_manager.login_user(email, password)
                    
                    if result.get("success"):
                        # Store session in session state
                        st.session_state.user_id = result['user_id']
                        st.session_state.user_email = result['email']
                        st.session_state.session_id = result['session_id']
                        st.session_state.authenticated = True
                        
                        st.success("✅ Signed in successfully!")
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(f"❌ {result.get('error', 'Login failed')}")
                else:
                    st.error("❌ Authentication system not available")
        
        st.markdown("---")
        st.markdown("### Don't have an account?")
        if st.button("📝 Sign Up", use_container_width=False, key="goto_signup"):
            st.session_state.show_signin = False
            st.rerun()


def show_subscription_page(user_id: str):
    """Display subscription page with coupon code input (shown after sign in)."""
    st.markdown("### 📦 Subscription Plans")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div style="border: 2px solid #e0e0e0; border-radius: 12px; padding: 2rem; text-align: center; background: #f8f9fa;">
            <h3 style="color: #000000;">Free</h3>
            <h2 style="color: #000000;">$0/month</h2>
            <ul style="text-align: left; color: #000000; padding-left: 1.5rem;">
                <li>10 extractions/month</li>
                <li>50MB max file size</li>
                <li>Basic features</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div style="border: 2px solid #1f77b4; border-radius: 12px; padding: 2rem; text-align: center; background: #e3f2fd;">
            <h3 style="color: #000000;">Pro</h3>
            <h2 style="color: #000000;">$9.99/month</h2>
            <ul style="text-align: left; color: #000000; padding-left: 1.5rem;">
                <li>100 extractions/month</li>
                <li>500MB max file size</li>
                <li>Priority processing</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div style="border: 2px solid #ff9800; border-radius: 12px; padding: 2rem; text-align: center; background: #fff3e0;">
            <h3 style="color: #000000;">Enterprise</h3>
            <h2 style="color: #000000;">$49.99/month</h2>
            <ul style="text-align: left; color: #000000; padding-left: 1.5rem;">
                <li>Unlimited extractions</li>
                <li>Unlimited file size</li>
                <li>API access</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Check if user is admin - auto-grant enterprise access
    admin_email = os.getenv('ADMIN_EMAIL', '').lower()
    try:
        if hasattr(st, 'secrets') and 'ADMIN_EMAIL' in st.secrets:
            admin_email = st.secrets['ADMIN_EMAIL'].lower()
    except:
        pass
    
    user_email = st.session_state.get('user_email', '').lower()
    is_admin = admin_email and user_email == admin_email
    
    if is_admin:
        st.success("🔑 **Admin Access Detected** - You have Enterprise access!")
        if st.button("🚀 Continue to App", type="primary", key="admin_continue"):
            st.session_state.subscription_active = True
            st.session_state.subscription_tier = 'enterprise'
            st.rerun()
        return
    
    # Coupon code section
    st.markdown("### 🎟️ Redeem Coupon Code")
    st.markdown("Enter a coupon code to get a subscription!")
    
    # Built-in coupons that work everywhere
    BUILTIN_COUPONS = {
        'ADMIN2024': {'tier': 'enterprise', 'days': 365},
        'FREETRIAL': {'tier': 'pro', 'days': 30},
        'WELCOME': {'tier': 'free', 'days': 365},
    }
    
    coupon_code = st.text_input(
        "Coupon Code",
        placeholder="Enter your coupon code here",
        help="Enter a valid coupon code to activate your subscription",
        key="coupon_input"
    )
    
    if st.button("🎁 Redeem Coupon", type="primary", use_container_width=False, key="redeem_coupon"):
        if not coupon_code:
            st.error("❌ Please enter a coupon code")
        else:
            # Check built-in coupons first
            coupon_upper = coupon_code.strip().upper()
            if coupon_upper in BUILTIN_COUPONS:
                coupon_info = BUILTIN_COUPONS[coupon_upper]
                st.success(f"✅ Coupon redeemed! You now have a {coupon_info['tier'].title()} subscription for {coupon_info['days']} days.")
                st.session_state.subscription_active = True
                st.session_state.subscription_tier = coupon_info['tier']
                st.rerun()
            elif SUBSCRIPTION_AVAILABLE:
                manager = SubscriptionManager()
                result = manager.apply_coupon_code(user_id, coupon_code)
                
                if result.get("success"):
                    st.success(f"✅ Coupon redeemed successfully! You now have a {result['tier']} subscription.")
                    st.info(f"📅 Subscription valid until: {result['end_date']}")
                    st.session_state.subscription_active = True
                    st.rerun()
                else:
                    st.error(f"❌ {result.get('error', 'Invalid coupon code')}")
            else:
                st.error("❌ Invalid coupon code")
    
    st.markdown("---")
    st.markdown("### 💡 Don't have a coupon?")
    st.info("💬 Contact support to get a free trial coupon code!")


def main():
    """Main application."""
    initialize_session_state()
    
    # Check authentication first
    if AUTH_AVAILABLE:
        auth_manager = AuthManager()
        
        # Check if user is authenticated
        if not st.session_state.get('authenticated', False):
            # Check if there's a valid session
            session_id = st.session_state.get('session_id')
            if session_id:
                user_info = auth_manager.verify_session(session_id)
                if user_info:
                    st.session_state.authenticated = True
                    st.session_state.user_id = user_info['user_id']
                    st.session_state.user_email = user_info['email']
                else:
                    # Session expired or invalid
                    st.session_state.authenticated = False
                    st.session_state.session_id = None
        
        # Show sign up/sign in if not authenticated
        if not st.session_state.get('authenticated', False):
            # Determine which page to show
            show_signin = st.session_state.get('show_signin', True)
            
            if show_signin:
                show_signin_page()
            else:
                show_signup_page()
            st.stop()
        
        # User is authenticated - get user_id
        user_id = st.session_state.get('user_id')
        user_email = st.session_state.get('user_email', '')
        
        # Initialize subscription and security
        if SUBSCRIPTION_AVAILABLE:
            subscription_manager = SubscriptionManager()
            security_manager = SecurityManager()
            
            # Get user subscription
            subscription = subscription_manager.get_user_subscription(user_id)
            
            # Check if admin bypass is active (from session state)
            if st.session_state.get('subscription_active'):
                tier = st.session_state.get('subscription_tier', 'enterprise')
                subscription = {
                    'tier': tier,
                    'status': 'active',
                    'admin_bypass': True,
                    'tier_info': {
                        'name': tier.title(),
                        'price': 0,
                        'max_extractions_per_month': -1,  # Unlimited
                        'max_file_size_mb': -1,  # Unlimited
                        'features': ['All features', 'Admin access', 'Unlimited usage']
                    }
                }
            
            # Check if user needs to subscribe
            if not subscription:
                # Show subscription page (user is signed in)
                st.markdown("""
                <div style="text-align: center; padding: 2rem 0 1rem 0; background-color: #ffffff;">
                    <div style="font-size: 5rem; margin-bottom: 0.5rem;">✨</div>
                    <h1 style="
                        font-size: 4.5rem;
                        font-weight: 900;
                        color: #000000 !important;
                        margin: 0;
                        letter-spacing: -0.03em;
                        line-height: 1;
                    ">ReqIQ</h1>
                    <p style="
                        font-size: 1.5rem;
                        color: #333333 !important;
                        margin: 0.75rem 0 0 0;
                        font-weight: 500;
                    ">Welcome, {email}!</p>
                    <p style="
                        font-size: 1rem;
                        color: #666666 !important;
                        margin: 0.5rem 0 0 0;
                    ">Get started with a subscription</p>
                </div>
                """.format(email=user_email), unsafe_allow_html=True)
                st.markdown("---")
                
                # Logout button
                col1, col2, col3 = st.columns([3, 1, 3])
                with col2:
                    if st.button("🚪 Sign Out", key="logout_button"):
                        session_id = st.session_state.get('session_id')
                        if session_id:
                            auth_manager.logout_user(session_id)
                        # Clear session state
                        for key in ['authenticated', 'user_id', 'user_email', 'session_id', 'subscription']:
                            if key in st.session_state:
                                del st.session_state[key]
                        st.rerun()
                
                show_subscription_page(user_id)
                st.stop()
            
            # Store subscription info in session state
            st.session_state.subscription = subscription
            st.session_state.user_id = user_id
            st.session_state.subscription_manager = subscription_manager
            st.session_state.security_manager = security_manager
    else:
        # Fallback if auth not available - use old method
        if SUBSCRIPTION_AVAILABLE:
            from subscription_manager import generate_user_id
            user_id = generate_user_id()
            subscription_manager = SubscriptionManager()
            security_manager = SecurityManager()
            
            subscription = subscription_manager.get_user_subscription(user_id)
            if not subscription:
                show_subscription_page(user_id)
                st.stop()
            
            st.session_state.subscription = subscription
            st.session_state.user_id = user_id
            st.session_state.subscription_manager = subscription_manager
            st.session_state.security_manager = security_manager
    
    # Enhanced Header with premium branding - Short & Catchy Name
    st.markdown("""
    <div style="text-align: center; padding: 2rem 0 1rem 0; background-color: #ffffff;">
        <div style="font-size: 5rem; margin-bottom: 0.5rem;">✨</div>
        <h1 style="
            font-size: 4.5rem;
            font-weight: 900;
            color: #000000 !important;
            margin: 0;
            letter-spacing: -0.03em;
            line-height: 1;
        ">ReqIQ</h1>
        <p style="
            font-size: 1.5rem;
            color: #333333 !important;
            margin: 0.75rem 0 0 0;
            font-weight: 500;
            letter-spacing: 0.02em;
        ">AI-Powered Requirements Extraction</p>
        <p style="
            font-size: 1rem;
            color: #555555 !important;
            margin: 0.5rem 0 0 0;
            font-style: italic;
            font-weight: 400;
        ">Transform meeting discussions into structured requirements</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    
    # Enhanced Sidebar for configuration
    with st.sidebar:
        # Add CSS for sidebar text visibility - BLACK text on LIGHT background (MAXIMUM PRIORITY)
        st.markdown("""
        <style>
        /* Sidebar Configuration section - LIGHT background with BLACK text - FORCE EVERYTHING */
        [data-testid="stSidebar"],
        section[data-testid="stSidebar"],
        .css-1d391kg,
        .css-1d391kg *,
        [data-testid="stSidebar"] *,
        section[data-testid="stSidebar"] * {
            background-color: #f8f9fa !important;
            color: #000000 !important;
        }
        
        /* Sidebar headers - BLACK text - MAXIMUM SPECIFICITY */
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] h4,
        [data-testid="stSidebar"] h5,
        [data-testid="stSidebar"] h6,
        section[data-testid="stSidebar"] h1,
        section[data-testid="stSidebar"] h2,
        section[data-testid="stSidebar"] h3,
        section[data-testid="stSidebar"] h4,
        section[data-testid="stSidebar"] h5,
        section[data-testid="stSidebar"] h6 {
            color: #000000 !important;
            background-color: transparent !important;
        }
        
        /* Sidebar text elements - BLACK text - FORCE ALL */
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] div,
        [data-testid="stSidebar"] span,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] strong,
        [data-testid="stSidebar"] em,
        [data-testid="stSidebar"] li,
        [data-testid="stSidebar"] ul,
        [data-testid="stSidebar"] ol,
        section[data-testid="stSidebar"] p,
        section[data-testid="stSidebar"] div,
        section[data-testid="stSidebar"] span,
        section[data-testid="stSidebar"] label {
            color: #000000 !important;
            background-color: transparent !important;
        }
        
        /* Sidebar info boxes and markdown - BLACK text */
        [data-testid="stSidebar"] .stInfo,
        [data-testid="stSidebar"] .stInfo *,
        [data-testid="stSidebar"] .stMarkdown,
        [data-testid="stSidebar"] .stMarkdown *,
        [data-testid="stSidebar"] .stSuccess,
        [data-testid="stSidebar"] .stSuccess *,
        [data-testid="stSidebar"] .stWarning,
        [data-testid="stSidebar"] .stWarning *,
        [data-testid="stSidebar"] .stError,
        [data-testid="stSidebar"] .stError * {
            color: #000000 !important;
            background-color: transparent !important;
        }
        
        /* Sidebar radio buttons and inputs - ensure visibility - MAXIMUM SPECIFICITY */
        [data-testid="stSidebar"] .stRadio,
        [data-testid="stSidebar"] .stRadio *,
        [data-testid="stSidebar"] .stRadio > div,
        [data-testid="stSidebar"] .stRadio > div > div,
        [data-testid="stSidebar"] .stRadio > div > div > label,
        [data-testid="stSidebar"] .stRadio label,
        [data-testid="stSidebar"] .stRadio span,
        [data-testid="stSidebar"] .stRadio p,
        [data-testid="stSidebar"] .stRadio div,
        [data-testid="stSidebar"] .stTextInput,
        [data-testid="stSidebar"] .stTextInput *,
        [data-testid="stSidebar"] .stTextInput label,
        [data-testid="stSidebar"] .stSelectbox,
        [data-testid="stSidebar"] .stSelectbox *,
        [data-testid="stSidebar"] .stSelectbox label,
        [data-testid="stSidebar"] .stCheckbox,
        [data-testid="stSidebar"] .stCheckbox *,
        [data-testid="stSidebar"] .stCheckbox label {
            color: #000000 !important;
            background-color: transparent !important;
        }
        
        /* Radio button labels - force black text */
        [data-testid="stSidebar"] .stRadio > div > div > label > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div > div {
            color: #000000 !important;
        }
        
        /* Radio button text content */
        [data-testid="stSidebar"] [data-baseweb="radio"] label,
        [data-testid="stSidebar"] [data-baseweb="radio"] label *,
        [data-testid="stSidebar"] [data-baseweb="radio"] span {
            color: #000000 !important;
        }
        
        /* Sidebar buttons - visible with proper styling */
        [data-testid="stSidebar"] .stButton,
        [data-testid="stSidebar"] .stButton > button,
        [data-testid="stSidebar"] .stButton > button *,
        [data-testid="stSidebar"] button,
        [data-testid="stSidebar"] button *,
        [data-testid="stSidebar"] [data-baseweb="button"],
        [data-testid="stSidebar"] [data-baseweb="button"] * {
            color: #ffffff !important;
            background-color: #1f77b4 !important;
            border: 1px solid #1f77b4 !important;
        }
        
        /* Button hover state */
        [data-testid="stSidebar"] .stButton > button:hover,
        [data-testid="stSidebar"] button:hover {
            background-color: #1565a0 !important;
            border-color: #1565a0 !important;
        }
        
        /* Radio buttons - ULTRA SPECIFIC targeting for visibility - MAXIMUM PRIORITY */
        [data-testid="stSidebar"] .stRadio,
        [data-testid="stSidebar"] .stRadio *,
        [data-testid="stSidebar"] .stRadio > div,
        [data-testid="stSidebar"] .stRadio > div *,
        [data-testid="stSidebar"] .stRadio > div > div,
        [data-testid="stSidebar"] .stRadio > div > div *,
        [data-testid="stSidebar"] .stRadio > div > div > label,
        [data-testid="stSidebar"] .stRadio > div > div > label *,
        [data-testid="stSidebar"] .stRadio > div > div > label > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div *,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div *,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div *,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div > div,
        [data-testid="stSidebar"] .stRadio > div > div > label > div > div > div > div *,
        [data-testid="stSidebar"] .stRadio label,
        [data-testid="stSidebar"] .stRadio label *,
        [data-testid="stSidebar"] .stRadio span,
        [data-testid="stSidebar"] .stRadio p,
        [data-testid="stSidebar"] .stRadio div,
        [data-testid="stSidebar"] [data-baseweb="radio"],
        [data-testid="stSidebar"] [data-baseweb="radio"] *,
        [data-testid="stSidebar"] [data-baseweb="radio"] label,
        [data-testid="stSidebar"] [data-baseweb="radio"] label *,
        [data-testid="stSidebar"] [data-baseweb="radio"] span,
        [data-testid="stSidebar"] [data-baseweb="radio"] div,
        [data-testid="stSidebar"] [data-baseweb="radio"] p {
            color: #000000 !important;
            background-color: transparent !important;
        }
        
        /* Radio button circles/indicators - ensure they're visible */
        [data-testid="stSidebar"] .stRadio input[type="radio"],
        [data-testid="stSidebar"] [data-baseweb="radio"] input[type="radio"],
        [data-testid="stSidebar"] input[type="radio"] {
            border-color: #000000 !important;
            background-color: #ffffff !important;
        }
        
        /* Universal selector for sidebar radio buttons - catch everything */
        [data-testid="stSidebar"] .stRadio,
        [data-testid="stSidebar"] .stRadio * {
            color: #000000 !important;
        }
        
        /* Sidebar section headers and all markdown content */
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] h4,
        [data-testid="stSidebar"] h5,
        [data-testid="stSidebar"] h6 {
            color: #000000 !important;
        }
        
        /* Override any Streamlit default dark mode styles */
        [data-testid="stSidebar"] [class*="css"],
        [data-testid="stSidebar"] [class*="st"] {
            color: #000000 !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown("### ⚙️ Configuration")
        st.markdown("---")
        
        # AGGRESSIVE fix for radio button visibility - runs continuously
        st.markdown("""
        <script>
        // Ultra-aggressive function that runs continuously using requestAnimationFrame
        function forceRadioButtonText() {
            const sidebar = document.querySelector('[data-testid="stSidebar"]');
            if (!sidebar) {
                requestAnimationFrame(forceRadioButtonText);
                return;
            }
            
            // Find ALL possible radio button containers
            const selectors = [
                '.stRadio',
                '[data-baseweb="radio"]',
                '[class*="radio"]',
                '[class*="Radio"]'
            ];
            
            let foundAny = false;
            selectors.forEach(function(selector) {
                try {
                    const radios = sidebar.querySelectorAll(selector);
                    radios.forEach(function(radio) {
                        foundAny = true;
                        // Force black on container
                        radio.style.setProperty('color', '#000000', 'important');
                        
                        // Get ALL descendants - use getElementsByTagName for maximum coverage
                        const allElements = radio.getElementsByTagName('*');
                        for (let i = 0; i < allElements.length; i++) {
                            const el = allElements[i];
                            // Skip radio input elements
                            if (el.tagName === 'INPUT' && el.type === 'radio') {
                                continue;
                            }
                            // Force black text on EVERYTHING else
                            el.style.setProperty('color', '#000000', 'important');
                            el.style.setProperty('-webkit-text-fill-color', '#000000', 'important');
                        }
                        
                        // Also handle text nodes by setting parent color
                        const walker = document.createTreeWalker(
                            radio,
                            NodeFilter.SHOW_TEXT,
                            null,
                            false
                        );
                        let textNode;
                        while (textNode = walker.nextNode()) {
                            if (textNode.parentElement) {
                                textNode.parentElement.style.setProperty('color', '#000000', 'important');
                            }
                        }
                    });
                } catch(e) {
                    // Ignore selector errors
                }
            });
            
            // Continue running
            requestAnimationFrame(forceRadioButtonText);
        }
        
        // Start immediately
        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', function() {
                requestAnimationFrame(forceRadioButtonText);
            });
        } else {
            requestAnimationFrame(forceRadioButtonText);
        }
        
        // Also use MutationObserver as backup
        const observer = new MutationObserver(function() {
            // Trigger on any change
        });
        
        const sidebar = document.querySelector('[data-testid="stSidebar"]');
        if (sidebar) {
            observer.observe(sidebar, {
                childList: true,
                subtree: true,
                attributes: true,
                attributeFilter: ['style', 'class']
            });
        }
        </script>
        """, unsafe_allow_html=True)
        
        # Transcription method selection
        transcription_method = st.radio(
            "Transcription Method",
            options=["Local Whisper (No API Key)", "OpenAI API (Requires Key)"],
            index=0 if WHISPER_LOCAL_AVAILABLE else 1,
            help="Local Whisper runs on your computer (free, no API key). OpenAI API is faster but requires an API key."
        )
        st.session_state.use_local_whisper = (transcription_method == "Local Whisper (No API Key)")
        
        # Requirements extraction method
        extraction_method = st.radio(
            "Requirements Extraction Method",
            options=["Ollama (Local, No API Key)", "OpenAI API (Requires Key)"],
            index=0,
            help="Ollama runs locally (free, no API key). OpenAI API is faster but requires an API key."
        )
        st.session_state.use_ollama = (extraction_method == "Ollama (Local, No API Key)")
        
        # Add another JavaScript after both radios are created
        st.markdown("""
        <script>
        setTimeout(function() {
            const sidebar = document.querySelector('[data-testid="stSidebar"]');
            if (sidebar) {
                const radios = sidebar.querySelectorAll('.stRadio, [data-baseweb="radio"]');
                radios.forEach(function(radio) {
                    const all = radio.querySelectorAll('*');
                    all.forEach(function(el) {
                        if (el.tagName !== 'INPUT' || el.type !== 'radio') {
                            el.style.color = '#000000';
                            el.style.setProperty('color', '#000000', 'important');
                        }
                    });
                });
            }
        }, 500);
        </script>
        """, unsafe_allow_html=True)
        
        # API Key input (only needed for OpenAI)
        if st.session_state.use_ollama:
            st.info("✅ Using Ollama for requirements extraction - no API key needed!")
            st.session_state.api_key = None
            
            # Ollama model selection
            st.markdown("""
            <style>
            /* Force black text in Ollama model input */
            .stTextInput input[value*="llama"] {
                color: #000000 !important;
            }
            </style>
            """, unsafe_allow_html=True)
            ollama_model = st.text_input(
                "Ollama Model Name",
                value=st.session_state.get('ollama_model', 'llama3.2'),
                help="Name of the Ollama model to use (e.g., llama3.2, mistral, codellama). Make sure you've pulled it: ollama pull <model-name>"
            )
            st.session_state.ollama_model = ollama_model
        else:
            if st.session_state.use_local_whisper:
                st.info("💡 Using local Whisper for transcription (no API key needed for transcription)")
            
            # Show API key source if auto-loaded
            api_key_source = st.session_state.get('api_key_source', '')
            if api_key_source == 'auto_loaded' and st.session_state.get('api_key'):
                st.success("✅ API key loaded automatically (from environment/config)")
            
            # Get current API key value
            current_key = st.session_state.get('api_key', '') or get_api_key_from_all_sources() or ''
            
            # Add inline style to force black text
            st.markdown("""
            <style>
            /* Force black text in API key input - maximum specificity */
            input[type="password"],
            input[type="text"] {
                color: #000000 !important;
                -webkit-text-fill-color: #000000 !important;
            }
            </style>
            """, unsafe_allow_html=True)
            
            api_key_input = st.text_input(
                "OpenAI API Key (for Requirements Extraction)" if st.session_state.use_local_whisper else "OpenAI API Key",
                type="password",
                help="Enter your OpenAI API key. It will be saved if you check 'Remember API Key' below.",
                value=current_key
            )
            
            # Remember API Key checkbox
            remember_key = st.checkbox(
                "💾 Remember API Key",
                value=False,
                help="Save API key to local config file for future sessions. Key is encrypted and stored securely."
            )
            
            # Update session state
            if api_key_input:
                st.session_state.api_key = api_key_input
                # Save to local config if checkbox is checked
                if remember_key:
                    if save_api_key(api_key_input):
                        st.success("✅ API key saved! It will be loaded automatically next time.")
                        st.session_state.api_key_source = "saved"
                    else:
                        st.warning("⚠️ Could not save API key. Check permissions or use environment variable instead.")
            else:
                # Try to get from auto-sources
                auto_key = get_api_key_from_all_sources()
                if auto_key:
                    st.session_state.api_key = auto_key
                else:
                    st.session_state.api_key = None
            
            # Show option to clear saved key
            if load_api_key():
                if st.button("🗑️ Clear Saved API Key"):
                    config_file = Path.home() / ".reqiq_config"
                    if config_file.exists():
                        config_file.unlink()
                        st.session_state.api_key = None
                        st.session_state.api_key_source = None
                        st.success("✅ Saved API key cleared!")
                        st.rerun()
            
            # Model selection (only for OpenAI)
            st.markdown("""
            <style>
            /* Force black text in selectbox */
            .stSelectbox select,
            .stSelectbox input {
                color: #000000 !important;
                -webkit-text-fill-color: #000000 !important;
            }
            </style>
            """, unsafe_allow_html=True)
            model = st.selectbox(
                "OpenAI Model",
                options=["gpt-4o-mini", "gpt-4o", "gpt-3.5-turbo"],
                index=0,
                help="Select the OpenAI model to use for extraction"
            )
            st.session_state.model = model
        
        st.markdown("---")
        
        # User info and logout
        if AUTH_AVAILABLE and st.session_state.get('authenticated'):
            user_email = st.session_state.get('user_email', '')
            st.markdown(f"### 👤 Signed in as: {user_email}")
            if st.button("🚪 Sign Out", key="sidebar_logout"):
                session_id = st.session_state.get('session_id')
                if session_id:
                    auth_manager = AuthManager()
                    auth_manager.logout_user(session_id)
                # Clear session state
                for key in ['authenticated', 'user_id', 'user_email', 'session_id', 'subscription']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
            st.markdown("---")
        
        # Subscription status display
        if SUBSCRIPTION_AVAILABLE and 'subscription' in st.session_state:
            subscription = st.session_state.subscription
            subscription_manager = st.session_state.subscription_manager
            user_id = st.session_state.user_id
            tier_info = subscription['tier_info']
            usage_info = subscription_manager.check_usage_limit(user_id, "extraction")
            
            st.markdown("### 💳 Subscription Status")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Plan", subscription['tier'].upper())
            with col2:
                if usage_info['limit'] == 'unlimited':
                    st.metric("Usage", f"{usage_info['current']} (Unlimited)")
                else:
                    st.metric("Usage", f"{usage_info['current']}/{usage_info['limit']}")
            with col3:
                if usage_info['limit'] != 'unlimited':
                    st.metric("Remaining", usage_info['remaining'])
            
            if subscription.get('end_date'):
                from datetime import datetime
                end_date = datetime.fromisoformat(subscription['end_date'])
                days_left = (end_date - datetime.now()).days
                if days_left > 0:
                    st.info(f"📅 Subscription expires in {days_left} days")
                else:
                    st.warning("⚠️ Subscription expired. Please renew.")
        
        st.markdown("---")
        
        # Instructions with better formatting
        st.markdown("### 📖 How to Use")
        st.markdown("""
        <div style="background: #f8f9fa; padding: 1rem; border-radius: 8px; margin: 1rem 0;">
        <ol style="margin: 0; padding-left: 1.5rem;">
        <li style="margin-bottom: 0.5rem;"><strong>Upload</strong> a file or paste text</li>
        <li style="margin-bottom: 0.5rem;"><strong>Configure</strong> API key and model</li>
        <li style="margin-bottom: 0.5rem;"><strong>Extract</strong> requirements</li>
        <li><strong>Download</strong> results</li>
        </ol>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("### 📁 Supported Formats")
        st.markdown("""
        <div style="background: #f8f9fa; padding: 1rem; border-radius: 8px;">
        <p style="margin: 0.5rem 0;">📄 Text files (.txt)</p>
        <p style="margin: 0.5rem 0;">📝 WebVTT files (.vtt)</p>
        <p style="margin: 0.5rem 0;">📊 JSON files (.json)</p>
        <p style="margin: 0.5rem 0;">🎬 Video files (.mp4, .mov, .avi, .mkv, .webm)</p>
        <p style="margin: 0.5rem 0;">🎤 Audio files (.m4a, .mp3, .wav, .flac, .ogg, .aac, .wma)</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Example format with better styling
        st.markdown("### 📝 Example Format")
        with st.expander("View example", expanded=False):
            st.code("""
Speaker Name: Message text here
Another Speaker: Another message
            """, language="text")
    
    # Enhanced Main content area with better tabs
    tab1, tab2 = st.tabs(["📁 Upload File", "📝 Paste Text"])
    
    with tab1:
        st.markdown("### 📤 Upload Your File")
        st.markdown("Upload a transcript file or video recording from your Teams meeting")
        
        # Add custom CSS for file uploader text visibility - BLACK text on LIGHT background
        st.markdown("""
        <style>
        /* File uploader - LIGHT background with BLACK text */
        [data-testid="stFileUploader"],
        [data-testid="stFileUploader"] > div,
        [data-testid="stFileUploader"] > div > div,
        [data-testid="stFileUploader"] > div > div > div,
        [data-testid="stFileUploader"] > div > div > div > div {
            color: #000000 !important;
            background-color: #f8f9fa !important;
            border: 2px dashed #cccccc !important;
        }
        
        /* File uploader text - BLACK */
        .stFileUploader label,
        .stFileUploader p,
        .stFileUploader div,
        .stFileUploader span,
        [data-testid="stFileUploader"] label,
        [data-testid="stFileUploader"] p,
        [data-testid="stFileUploader"] div,
        [data-testid="stFileUploader"] span,
        [data-testid="stFileUploader"] label p,
        [data-testid="stFileUploader"] label div,
        [data-testid="stFileUploader"] > div > div > div,
        [data-testid="stFileUploader"] > div > div > div > div,
        [data-testid="stFileUploader"] > div > div > div > div > div,
        [data-testid="stFileUploader"] > div > div > div > div > div > div,
        [data-testid="stFileUploader"] > div > div > div > div > div > div > div,
        [data-testid="stFileUploader"] > div > div > div > div > div > div > div > div {
            color: #000000 !important;
            background-color: transparent !important;
        }
        
        /* File uploader drop zone - light background */
        [data-testid="stFileUploader"] {
            color: #000000 !important;
            background-color: #f8f9fa !important;
        }
        
        /* File uploader text elements - all BLACK */
        .stFileUploader *,
        [data-testid="stFileUploader"] * {
            color: #000000 !important;
            background-color: transparent !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader(
            "Choose a transcript file, video file, or audio file",
            type=['txt', 'vtt', 'json', 'mp4', 'mov', 'avi', 'mkv', 'webm', 'm4a', 'mp3', 'wav', 'flac', 'ogg', 'wma', 'aac'],
            help="Upload a transcript file, video recording, or audio file (MP4, MOV, AVI, MKV, M4A, MP3, WAV, FLAC, etc.)",
            label_visibility="collapsed"
        )
        
        if uploaded_file is not None:
            # Security: Validate file upload
            if SUBSCRIPTION_AVAILABLE and 'security_manager' in st.session_state:
                security_manager = st.session_state.security_manager
                subscription = st.session_state.get('subscription')
                max_size_mb = subscription['tier_info']['max_file_size_mb'] if subscription else 50
                
                validation = security_manager.validate_file_upload(uploaded_file, max_size_mb)
                if not validation['valid']:
                    st.error("❌ File validation failed:")
                    for error in validation['errors']:
                        st.error(f"  • {error}")
                    uploaded_file = None
                    st.stop()
            
            file_extension = Path(uploaded_file.name).suffix.lower()
            video_formats = ['.mp4', '.mov', '.avi', '.mkv', '.webm']
            audio_formats = ['.m4a', '.mp3', '.wav', '.flac', '.ogg', '.aac', '.wma']
            is_video = file_extension in video_formats
            is_audio = file_extension in audio_formats
            
            if is_video or is_audio:
                # Enhanced video/audio file display
                col1, col2 = st.columns([3, 1])
                with col1:
                    file_type = "🎬 Video" if is_video else "🎤 Audio"
                    st.success(f"✅ **{uploaded_file.name}** ({file_type})")
                with col2:
                    file_size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
                    st.metric("Size", f"{file_size_mb:.2f} MB")
                
                st.info("💡 Large files will be automatically chunked for processing.")
            else:
                # Enhanced text file display
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.success(f"✅ **{uploaded_file.name}**")
                with col2:
                    file_size_kb = len(uploaded_file.getvalue()) / 1024
                    st.metric("Size", f"{file_size_kb:.2f} KB")
                
                # Show file preview for text files with better styling
                with st.expander("📄 Preview File Content", expanded=False):
                    try:
                        content = uploaded_file.read().decode('utf-8')
                        st.text_area("File content", content, height=200, disabled=True, label_visibility="collapsed")
                        uploaded_file.seek(0)  # Reset file pointer
                    except:
                        st.info("Preview not available for this file type")
                        uploaded_file.seek(0)  # Reset file pointer
    
    with tab2:
        st.markdown("### 📝 Paste Transcript Text")
        st.markdown("Paste your meeting transcript directly into the text area below")
        
        # Force black text in textarea
        st.markdown("""
        <style>
        /* Force black text in textarea - maximum specificity */
        textarea {
            color: #000000 !important;
            -webkit-text-fill-color: #000000 !important;
        }
        .stTextArea textarea {
            color: #000000 !important;
            -webkit-text-fill-color: #000000 !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        transcript_text = st.text_area(
            "Enter transcript text",
            height=300,
            placeholder="Speaker Name: Message text here\nAnother Speaker: Another message\n\nExample:\nJohn Doe: We need to implement user authentication\nJane Smith: That's a good point. We should also add 2FA.",
            help="Paste your transcript text here. Use format: 'Speaker: message' or '[Speaker] message'",
            label_visibility="collapsed"
        )
        st.session_state.transcript_text = transcript_text
    
    # Enhanced Extract button with better positioning
    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        extract_button = st.button(
            "🚀 Extract Requirements",
            type="primary",
            use_container_width=False,
            disabled=(uploaded_file is None and not st.session_state.transcript_text),
            help="Click to start extracting requirements from your transcript"
        )
        if uploaded_file is None and not st.session_state.transcript_text:
            st.caption("👆 Upload a file or paste text to enable extraction")
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Processing
    if extract_button:
        # Security: Check subscription and usage limits
        if SUBSCRIPTION_AVAILABLE:
            user_id = st.session_state.get('user_id')
            subscription_manager = st.session_state.get('subscription_manager')
            
            if user_id and subscription_manager:
                usage_check = subscription_manager.check_usage_limit(user_id, "extraction")
                if not usage_check['allowed']:
                    st.error(f"❌ Usage limit exceeded! You've used {usage_check['current']} out of {usage_check['limit']} extractions this month.")
                    st.info("💡 Upgrade your subscription or wait for next month's reset.")
                    st.stop()
        
        # Get settings
        use_local = st.session_state.get('use_local_whisper', False)
        use_ollama = st.session_state.get('use_ollama', False)
        api_key = None
        
        # API key only needed if using OpenAI for extraction
        if not use_ollama:
            api_key = st.session_state.api_key or os.getenv('OPENAI_API_KEY')
            if not api_key:
                if use_local:
                    st.error("❌ Please provide an OpenAI API key in the sidebar for requirements extraction, or switch to Ollama (Local, No API Key).")
                else:
                    st.error("❌ Please provide an OpenAI API key in the sidebar or set OPENAI_API_KEY environment variable, or switch to Ollama (Local, No API Key).")
                st.stop()
        
        # Get model settings
        if use_ollama:
            ollama_model = st.session_state.get('ollama_model', 'llama3.2')
            model = None  # Not used for Ollama
        else:
            model = st.session_state.get('model', 'gpt-4o-mini')
            ollama_model = None
        
        # Check if it's a video or audio file
        is_media_file = False
        is_video_file = False
        is_audio_file = False
        if uploaded_file:
            file_extension = Path(uploaded_file.name).suffix.lower()
            video_formats = ['.mp4', '.mov', '.avi', '.mkv', '.webm']
            audio_formats = ['.m4a', '.mp3', '.wav', '.flac', '.ogg', '.aac', '.wma']
            is_video_file = file_extension in video_formats
            is_audio_file = file_extension in audio_formats
            is_media_file = is_video_file or is_audio_file
        
        # Process video, audio, or transcript
        if is_media_file:
            # Process video or audio file
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                use_local = st.session_state.get('use_local_whisper', False)
                
                if is_video_file:
                    if not MOVIEPY_AVAILABLE:
                        st.error("❌ moviepy is required for video processing. Please install it: `pip3 install moviepy`")
                        st.stop()
                    status_text.info("🎬 Processing video file...")
                    messages, error = process_video_file(uploaded_file, api_key, progress_bar, status_text, use_local=use_local)
                else:
                    # Process audio file (m4a, mp3, wav, etc.)
                    status_text.info("🎤 Processing audio file...")
                    messages, error = process_audio_file(uploaded_file, api_key, progress_bar, status_text, use_local=use_local)
                
                if error:
                    media_type = "video" if is_video_file else "audio"
                    st.error(f"❌ Error processing {media_type} file: {error}")
                    # Show full error details in an expander
                    with st.expander("🔍 View Full Error Details (for debugging)"):
                        st.code(error, language="text")
                        st.info("💡 Copy this error message and share it for debugging")
                    st.stop()
                
                if not messages:
                    media_type = "video" if is_video_file else "audio"
                    st.warning(f"⚠️ No transcript generated from {media_type} file")
                    st.stop()
                
                progress_bar.progress(1.0)
                media_type = "video" if is_video_file else "audio"
                status_text.success(f"✅ Generated transcript from {media_type} file with {len(messages)} segments")
                st.session_state.messages_parsed = messages
                
            except Exception as e:
                error_msg = str(e)
                st.error(f"❌ Error: {error_msg}")
                # Show full error details in an expander
                with st.expander("🔍 View Full Error Details (for debugging)"):
                    st.code(error_msg, language="text")
                    st.info("💡 Copy this error message and share it for debugging")
                st.stop()
            finally:
                # Keep progress bar visible briefly, then clear
                import time
                time.sleep(0.5)
                progress_bar.empty()
                status_text.empty()
        
        else:
            # Parse transcript
            with st.spinner("📖 Parsing transcript..."):
                if uploaded_file:
                    messages, error = parse_transcript_file(uploaded_file)
                elif st.session_state.transcript_text:
                    messages, error = parse_transcript_text(st.session_state.transcript_text)
                else:
                    st.error("❌ Please upload a file or paste transcript text")
                    st.stop()
                
                if error:
                    st.error(f"❌ Error parsing transcript: {error}")
                    st.stop()
                
                if not messages:
                    st.warning("⚠️ No messages found in transcript")
                    st.stop()
                
                st.session_state.messages_parsed = messages
                st.success(f"✅ Parsed {len(messages)} messages")
        
        # Extract requirements (incrementally for large transcripts)
        extraction_method_name = "Ollama (local)" if use_ollama else f"OpenAI ({model})"
        
        # Create progress tracking
        extraction_progress = st.progress(0)
        extraction_status = st.empty()
        results_container = st.container()
        
        def extraction_progress_callback(progress, message):
            """Update progress and show partial results."""
            extraction_progress.progress(progress)
            extraction_status.info(f"📊 {message}")
            
            # Show partial results if available
            if st.session_state.get('partial_requirements'):
                with results_container:
                    st.markdown("### 📋 Partial Results (Updated as chunks are processed)")
                    latest_chunk = st.session_state.partial_requirements[-1]
                    st.info(f"✅ Processed chunk {latest_chunk['chunk']}/{latest_chunk['total_chunks']}")
                    
                    # Show summary of latest chunk
                    chunk_req = latest_chunk['requirements']
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("FR", len(chunk_req.get('functional_requirements', [])))
                    with col2:
                        st.metric("NFR", len(chunk_req.get('non_functional_requirements', [])))
                    with col3:
                        st.metric("Action Items", len(chunk_req.get('action_items', [])))
                    with col4:
                        st.metric("Decisions", len(chunk_req.get('decisions', [])))
        
        # Clear previous partial results
        st.session_state.partial_requirements = []
        
        try:
            # Determine chunk size based on message count
            chunk_size = 50 if len(messages) > 100 else len(messages)
            
            requirements, error = extract_requirements(
                messages, 
                api_key, 
                model,
                use_ollama=use_ollama,
                ollama_model=ollama_model if use_ollama else None,
                chunk_size=chunk_size,
                progress_callback=extraction_progress_callback
            )
            
            if error:
                st.error(f"❌ Error extracting requirements: {error}")
                extraction_progress.empty()
                extraction_status.empty()
                st.stop()
            
            st.session_state.requirements = requirements
            extraction_progress.progress(1.0)
            extraction_status.success("✅ Requirements extracted successfully!")
            
            # Track usage
            if SUBSCRIPTION_AVAILABLE and 'subscription_manager' in st.session_state:
                user_id = st.session_state.get('user_id')
                if user_id:
                    subscription_manager = st.session_state.subscription_manager
                    subscription_manager.track_usage(user_id, "extraction", {
                        "messages_count": len(messages),
                        "timestamp": datetime.now().isoformat()
                    })
            
            # Clear progress indicators after a moment
            import time
            time.sleep(1)
            extraction_progress.empty()
            extraction_status.empty()
            
        except Exception as e:
            error_msg = str(e)
            
            # Check if it's an Ollama-related error
            if "Ollama is not running" in error_msg or "Ollama" in error_msg:
                st.error("❌ Ollama is not running or not installed")
                with st.expander("📖 How to Install and Start Ollama"):
                    st.markdown("""
                    **Option 1: Install Ollama (Recommended for no API key)**
                    
                    1. **Install Ollama:**
                       ```bash
                       brew install ollama
                       ```
                       Or download from: https://ollama.ai
                    
                    2. **Start Ollama service:**
                       ```bash
                       ollama serve
                       ```
                       (It may start automatically after installation)
                    
                    3. **Pull a model:**
                       ```bash
                       ollama pull llama3.2
                       ```
                       (This downloads the model - may take a few minutes)
                    
                    **Option 2: Use OpenAI API instead**
                    
                    If you have an OpenAI API key, you can switch to "OpenAI API (Requires Key)" 
                    in the sidebar instead of using Ollama.
                    """)
            else:
                st.error(f"❌ Error extracting requirements: {error_msg}")
            
            extraction_progress.empty()
            extraction_status.empty()
            st.stop()
    
    # Display partial results if processing
    if st.session_state.get('partial_requirements') and not st.session_state.requirements:
        st.divider()
        st.header("📊 Processing in Progress...")
        st.info("🔄 Requirements are being extracted incrementally. Results will appear below as chunks are processed.")
        
        # Show latest chunk results
        if st.session_state.partial_requirements:
            latest = st.session_state.partial_requirements[-1]
            st.subheader(f"Latest Chunk ({latest['chunk']}/{latest['total_chunks']})")
            req = latest['requirements']
            
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            with col1:
                st.metric("Functional", len(req.get('functional_requirements', [])))
            with col2:
                st.metric("Non-Functional", len(req.get('non_functional_requirements', [])))
            with col3:
                st.metric("Business Rules", len(req.get('business_rules', [])))
            with col4:
                st.metric("Action Items", len(req.get('action_items', [])))
            with col5:
                st.metric("Decisions", len(req.get('decisions', [])))
            with col6:
                st.metric("Stakeholders", len(req.get('stakeholders', [])))
    
    # Display final results with enhanced styling
    if st.session_state.requirements:
        st.markdown("---")
        st.markdown("""
        <div style="text-align: center; padding: 1rem 0; background-color: #ffffff;">
            <h2 style="
                font-size: 2.5rem;
                font-weight: 700;
                color: #000000 !important;
                margin: 0;
            ">📊 Extracted Requirements</h2>
        </div>
        """, unsafe_allow_html=True)
        
        # Enhanced Summary statistics with better cards
        req = st.session_state.requirements
        st.markdown("### 📈 Summary")
        
        col1, col2, col3, col4, col5, col6 = st.columns(6)
        
        with col1:
            st.metric(
                "Functional", 
                len(req.get('functional_requirements', [])),
                help="Functional requirements extracted"
            )
        with col2:
            st.metric(
                "Non-Functional", 
                len(req.get('non_functional_requirements', [])),
                help="Non-functional requirements extracted"
            )
        with col3:
            st.metric(
                "Business Rules", 
                len(req.get('business_rules', [])),
                help="Business rules identified"
            )
        with col4:
            st.metric(
                "Action Items", 
                len(req.get('action_items', [])),
                help="Action items found"
            )
        with col5:
            st.metric(
                "Decisions", 
                len(req.get('decisions', [])),
                help="Decisions made"
            )
        with col6:
            st.metric(
                "Stakeholders", 
                len(req.get('stakeholders', [])),
                help="Stakeholders identified"
            )
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # Enhanced Detailed sections with better tab styling
        st.markdown("### 📑 Detailed View")
        tabs = st.tabs([
            "📋 Functional",
            "⚙️ Non-Functional",
            "📜 Business Rules",
            "✅ Action Items",
            "🎯 Decisions",
            "👥 Stakeholders",
            "📄 Full Report"
        ])
        
        with tabs[0]:
            if req.get('functional_requirements'):
                st.markdown(f"**Found {len(req['functional_requirements'])} functional requirement(s)**")
                st.markdown("<br>", unsafe_allow_html=True)
                for i, fr in enumerate(req['functional_requirements'], 1):
                    with st.expander(f"**{fr.get('id', f'FR-{i:03d}')}** - {fr.get('description', 'N/A')[:60]}...", expanded=False):
                        col_desc, col_meta = st.columns([2, 1])
                        with col_desc:
                            st.markdown("#### 📝 Description")
                            st.write(fr.get('description', 'N/A'))
                        with col_meta:
                            priority = fr.get('priority', 'Not specified')
                            priority_color = {
                                'High': '🔴',
                                'Medium': '🟡',
                                'Low': '🟢'
                            }.get(priority, '⚪')
                            st.markdown("#### 📊 Metadata")
                            st.write(f"**Priority:** {priority_color} {priority}")
                            st.write(f"**Source:** 👤 {fr.get('speaker', 'Unknown')}")
                        if fr.get('context'):
                            st.markdown("---")
                            st.markdown("#### 💬 Context")
                            st.write(fr.get('context'))
            else:
                st.info("ℹ️ No functional requirements found in this transcript")
        
        with tabs[1]:
            if req.get('non_functional_requirements'):
                for i, nfr in enumerate(req['non_functional_requirements'], 1):
                    with st.expander(f"{nfr.get('id', f'NFR-{i:03d}')}: {nfr.get('description', 'N/A')[:50]}..."):
                        st.write("**Description:**", nfr.get('description', 'N/A'))
                        st.write("**Priority:**", nfr.get('priority', 'Not specified'))
                        st.write("**Source:**", nfr.get('speaker', 'Unknown'))
                        if nfr.get('context'):
                            st.write("**Context:**", nfr.get('context'))
            else:
                st.info("No non-functional requirements found")
        
        with tabs[2]:
            if req.get('business_rules'):
                for i, rule in enumerate(req['business_rules'], 1):
                    with st.expander(f"{rule.get('id', f'BR-{i:03d}')}: {rule.get('description', rule.get('rule', 'N/A'))[:50]}..."):
                        st.write("**Rule:**", rule.get('description', rule.get('rule', 'N/A')))
                        st.write("**Source:**", rule.get('speaker', 'Unknown'))
            else:
                st.info("No business rules found")
        
        with tabs[3]:
            if req.get('action_items'):
                # Table view
                import pandas as pd
                df = pd.DataFrame(req['action_items'])
                st.dataframe(df, use_container_width=False)
            else:
                st.info("No action items found")
        
        with tabs[4]:
            if req.get('decisions'):
                for i, decision in enumerate(req['decisions'], 1):
                    with st.expander(f"{decision.get('id', f'D-{i:03d}')}: {decision.get('decision', 'N/A')[:50]}..."):
                        st.write("**Decision:**", decision.get('decision', 'N/A'))
                        st.write("**Rationale:**", decision.get('rationale', 'N/A'))
                        st.write("**Decision Maker:**", decision.get('decision_maker', 'Unknown'))
            else:
                st.info("No decisions found")
        
        with tabs[5]:
            if req.get('stakeholders'):
                for stakeholder in req['stakeholders']:
                    with st.expander(stakeholder.get('name', 'Unknown')):
                        st.write("**Role:**", stakeholder.get('role', 'N/A'))
                        st.write("**Interests:**", stakeholder.get('interests', 'N/A'))
            else:
                st.info("No stakeholders identified")
        
        with tabs[6]:
            # Full markdown report
            formatter = RequirementsFormatter()
            markdown_report = formatter.format_markdown(st.session_state.requirements)
            st.markdown(markdown_report)
        
        # Enhanced Download section
        st.markdown("---")
        st.markdown("### 💾 Download Results")
        st.markdown("Export your extracted requirements in your preferred format")
        
        col1, col2, col3 = st.columns([1, 1, 1])
        # Ensure datetime is available (re-import if needed)
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
        
        with col1:
            # Markdown download
            formatter = RequirementsFormatter()
            markdown_content = formatter.format_markdown(st.session_state.requirements)
            st.download_button(
                label="📄 Markdown",
                data=markdown_content,
                file_name=f"requirements_{timestamp}.md",
                mime="text/markdown",
                use_container_width=False,
                help="Download as formatted Markdown document"
            )
        
        with col2:
            # PDF download
            try:
                pdf_content = generate_pdf(st.session_state.requirements)
                st.download_button(
                    label="📑 PDF",
                    data=pdf_content,
                    file_name=f"requirements_{timestamp}.pdf",
                    mime="application/pdf",
                    use_container_width=False,
                    help="Download as formatted PDF document"
                )
            except ImportError as e:
                st.error(f"❌ {str(e)}")
            except Exception as e:
                st.error(f"❌ Error generating PDF: {str(e)}")
        
        with col3:
            # Excel download
            try:
                excel_content = generate_excel(st.session_state.requirements)
                st.download_button(
                    label="📊 Excel",
                    data=excel_content,
                    file_name=f"requirements_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=False,
                    help="Download as Excel spreadsheet with multiple sheets"
                )
            except ImportError as e:
                st.error(f"❌ {str(e)}")
            except Exception as e:
                st.error(f"❌ Error generating Excel: {str(e)}")
        
        # Feedback and Rerun section
        st.markdown("---")
        st.markdown("### 💬 Provide Feedback & Refine Extraction")
        st.markdown("Add feedback, corrections, or guidance to improve the extraction. Click 'Rerun Extraction' to process again with your feedback.")
        
        # Initialize feedback in session state if not present
        if 'user_feedback' not in st.session_state:
            st.session_state.user_feedback = ""
        
        # Handle clear feedback action
        if 'clear_feedback' in st.session_state and st.session_state.clear_feedback:
            st.session_state.user_feedback = ""
            st.session_state.clear_feedback = False
        
        feedback_text = st.text_area(
            label="Feedback / Corrections / Guidance",
            value=st.session_state.user_feedback,
            placeholder="Example:\n- Focus more on supplier requirements\n- The PO number mentioned should be included\n- Add more details about non-functional requirements\n- Correct any terminology errors",
            help="Provide feedback, corrections, or specific guidance to improve the next extraction",
            height=150,
            key="user_feedback"
        )
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            if st.button("🔄 Rerun Extraction with Feedback", type="primary", use_container_width=False):
                if not feedback_text or not feedback_text.strip():
                    st.warning("⚠️ Please provide feedback before rerunning extraction.")
                else:
                    # Check if we have messages to rerun
                    if not st.session_state.get('messages_parsed'):
                        st.error("❌ No transcript available. Please upload a file or paste text first.")
                    else:
                        # Show progress
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        def progress_callback(progress, message):
                            progress_bar.progress(progress)
                            status_text.info(message)
                        
                        try:
                            progress_callback(0.1, "🔄 Rerunning extraction with your feedback...")
                            
                            # Get extraction settings
                            api_key = st.session_state.get('api_key')
                            model = st.session_state.get('model', 'gpt-4o-mini')
                            use_ollama = st.session_state.get('use_ollama', False)
                            ollama_model = st.session_state.get('ollama_model', 'llama3.2')
                            
                            # Extract requirements with feedback
                            requirements, error = extract_requirements(
                                st.session_state.messages_parsed,
                                api_key,
                                model,
                                use_ollama,
                                ollama_model,
                                chunk_size=50,
                                progress_callback=progress_callback,
                                feedback=feedback_text
                            )
                            
                            if error:
                                st.error(f"❌ Error: {error}")
                            else:
                                st.session_state.requirements = requirements
                                progress_callback(1.0, "✅ Extraction complete with feedback!")
                                st.success("✅ Requirements updated based on your feedback!")
                                st.rerun()
                                
                        except Exception as e:
                            st.error(f"❌ Error during extraction: {str(e)}")
                            status_text.empty()
        
        with col2:
            if st.button("🗑️ Clear Feedback", use_container_width=False):
                # Set flag to clear feedback on next rerun
                st.session_state.clear_feedback = True
                st.rerun()


# Streamlit automatically executes the script when run with 'streamlit run'
# Just call main() - Streamlit will handle the execution context
if __name__ == "__main__" or True:  # Always run in Streamlit
    try:
        main()
    except Exception as e:
        st.error(f"❌ Error loading application: {str(e)}")
        st.exception(e)
        st.info("Please check the terminal for more details or contact support.")

